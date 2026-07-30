"""
Fallback manual de captura de publicações — planilha exportada do Legal One.

É a terceira camada de captura, atrás do `GET /Updates` (primária). Quando a
API do L1 está fora do ar, o operador exporta as publicações na tela do Legal
One e sobe o arquivo aqui. Por não depender de rede nenhuma, é o único caminho
que continua funcionando com tudo fora.

REGRA CENTRAL DO DESENHO: a planilha NÃO grava em `publicacao_registros`.
Ela é convertida para o mesmo contrato de dicionário que o
`fetch_all_publications` do L1 devolve e entregue a
`create_and_run_search(prefetched_publications=...)`. Com isso, a dedup em duas
camadas, a detecção de publicação obsoleta e a "enxugada" antes da
classificação — tudo que já existe e é testado — valem de graça, sem
duplicação de lógica e sem risco de divergir do fluxo automático.

A coluna `Id` da planilha é o ID do PROCESSO no Legal One (`lawsuit_id`), não
o do andamento. Verificado contra o banco de produção em 30/07/2026: dos 1.237
IDs de uma extração real, 1.013 existiam no `lawsuit_cache` e em 1.013 de 1.013
o CNJ do cache era idêntico ao da planilha. A relação Id↔Pasta é 1:1 (1.238
para 1.238, sem cruzamento), e os CNJs que se repetem são pastas distintas
(apenso `/001` ou pasta duplicada).

Essa coluna é o que torna o método seguro: sem ela seria preciso adivinhar o
processo pelo CNJ, e 1.370 CNJs da base têm MAIS DE UM `lawsuit_id`. Errar aí
significa criar tarefa no processo errado.
"""
from __future__ import annotations

import hashlib
import io
import logging
import re
import unicodedata
import zlib
from datetime import date, datetime
from typing import Any, Optional

from sqlalchemy.orm import Session

from app.models.legal_one import LegalOneOffice
from app.models.publication_search import PublicationRecord

logger = logging.getLogger(__name__)


# O contrato do L1 para publicações de diário. Mantido igual ao que o
# fluxo automático grava, pra planilha e API produzirem registros idênticos.
ORIGIN_TYPE = "OfficialJournalsCrawler"
UPDATE_TYPE_ID = 5

# Só linhas desse tipo de andamento viram publicação. A extração do L1 pode
# trazer outros andamentos na mesma tela.
TIPO_PUBLICACAO = "publicacao"

# Teto de linhas por arquivo. Uma extração de um dia inteiro fica na casa de
# 1.200 linhas; 20 mil cobre uma janela larga com folga e evita que um arquivo
# errado (ex.: export da base inteira) derrube o worker.
MAX_LINHAS = 20_000


# ── Identificação das colunas ──────────────────────────────────────────
#
# O casamento é por NOME normalizado (sem acento, minúsculo, espaços
# colapsados), nunca por posição: o operador monta a extração na tela do L1 e
# a ordem das colunas muda conforme o que ele seleciona.

def _norm(texto: Any) -> str:
    """Normaliza para comparação: sem acento, minúsculo, espaço colapsado."""
    s = str(texto or "").strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"\s+", " ", s)
    return s.lower()


# Cada campo aceita mais de um rótulo — a tela do L1 varia o nome conforme a
# versão do relatório, e o operador às vezes renomeia o cabeçalho.
COLUNAS: dict[str, tuple[str, ...]] = {
    "lawsuit_id": ("id", "id do processo", "id processo"),
    "cnj": ("n do processo", "no do processo", "numero do processo", "processo"),
    "pasta": ("pasta",),
    "escritorio": ("escritorio responsavel", "escritorio"),
    "responsavel": ("responsavel principal",),
    "data": (
        "andamentos / data/hora",
        "andamentos / data",
        "data/hora",
        "data do andamento",
    ),
    "descricao": (
        "andamentos / descricao",
        "descricao do andamento",
        "descricao",
    ),
    "tipo": ("andamentos / tipo", "tipo do andamento"),
    "data_cadastro": ("data do cadastro", "data de cadastro"),
}

# Sem essas quatro não há publicação que se sustente.
OBRIGATORIAS = ("lawsuit_id", "cnj", "data", "descricao")


def _mapear_colunas(header: tuple) -> tuple[dict[str, int], list[str]]:
    """Devolve ({campo: índice}, [campos obrigatórios ausentes])."""
    achado: dict[str, int] = {}
    normalizado = [_norm(h) for h in header]
    for campo, rotulos in COLUNAS.items():
        for idx, nome in enumerate(normalizado):
            if nome in rotulos and campo not in achado:
                achado[campo] = idx
                break
    faltando = [c for c in OBRIGATORIAS if c not in achado]
    return achado, faltando


# ── Normalização de valores ────────────────────────────────────────────

def _so_digitos(valor: Any) -> str:
    return re.sub(r"\D", "", str(valor or ""))


def _data_publicacao(valor: Any) -> Optional[str]:
    """Converte a data da planilha pro formato que o L1 devolve.

    O fluxo automático grava `2026-07-28T00:00:00Z`; a planilha precisa gravar
    igual, senão a chave de dedup (lawsuit_id, publication_date) não casa entre
    as duas fontes e a mesma publicação entraria duas vezes.
    """
    if valor is None or str(valor).strip() == "":
        return None
    if isinstance(valor, datetime):
        return valor.strftime("%Y-%m-%dT00:00:00Z")
    if isinstance(valor, date):
        return valor.strftime("%Y-%m-%dT00:00:00Z")
    texto = str(valor).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y %H:%M:%S",
                "%d/%m/%Y %H:%M", "%d/%m/%Y"):
        try:
            return datetime.strptime(texto[:19], fmt).strftime("%Y-%m-%dT00:00:00Z")
        except ValueError:
            continue
    # Já veio ISO com fuso/Z — aproveita só a parte da data.
    m = re.match(r"(\d{4}-\d{2}-\d{2})", texto)
    return f"{m.group(1)}T00:00:00Z" if m else None


def _data_iso(valor: Any) -> Optional[str]:
    """Data de cadastro da pasta — alimenta a detecção de publicação obsoleta."""
    if valor is None or str(valor).strip() == "":
        return None
    if isinstance(valor, (datetime, date)):
        return valor.strftime("%Y-%m-%dT%H:%M:%SZ") if isinstance(valor, datetime) \
            else valor.strftime("%Y-%m-%dT00:00:00Z")
    texto = str(valor).strip()
    m = re.match(r"(\d{4}-\d{2}-\d{2})", texto)
    if m:
        return f"{m.group(1)}T00:00:00Z"
    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y"):
        try:
            return datetime.strptime(texto[:19], fmt).strftime("%Y-%m-%dT00:00:00Z")
        except ValueError:
            continue
    return None


# ── ID sintético do andamento ──────────────────────────────────────────
#
# `legal_one_update_id` é NOT NULL e UNIQUE, e a planilha não traz o ID do
# andamento (só o do processo). Geramos um ID NEGATIVO e determinístico:
#
#   - negativo  → nunca colide com ID real do L1, que é positivo (a menor
#                 ocorrência em produção é 1.356.393);
#   - determinístico → subir a MESMA planilha de novo gera os MESMOS IDs, e a
#                 dedup por `legal_one_update_id` descarta sozinha. É assim que
#                 o método fica idempotente sem precisar de coluna nova.

def _hash_base(lawsuit_id: int, publication_date: str, descricao: str) -> int:
    corpo = hashlib.sha1(descricao.strip().encode("utf-8")).hexdigest()[:16]
    bruto = f"{lawsuit_id}|{publication_date}|{corpo}"
    return zlib.crc32(bruto.encode("utf-8")) & 0x7FFF_FFFF


def _id_sintetico(base: int, passo: int) -> int:
    valor = (base + passo) & 0x7FFF_FFFF
    return -(valor or 1)


def atribuir_update_ids(db: Session, linhas: list[dict]) -> None:
    """Atribui `id` a cada linha, resolvendo colisão de hash sem perder linha.

    O espaço de 31 bits é grande, mas não infinito: acumulando uploads, duas
    publicações diferentes podem cair no mesmo hash. Se a gente ignorasse isso,
    a segunda seria tratada como duplicata exata e SUMIRIA em silêncio — o
    oposto do que esse fallback existe pra fazer.

    Regra: se o ID candidato já está ocupado pelo MESMO (processo, data), pode
    reusar — é a mesma publicação reenviada, e a dedup vai descartar como
    duplicata, que é o comportamento correto. Se está ocupado por um par
    DIFERENTE, é colisão de verdade e o ID anda até achar espaço livre.
    """
    ocupados: dict[int, tuple[Optional[int], Optional[str]]] = {
        rid: (lid, pdate)
        for rid, lid, pdate in db.query(
            PublicationRecord.legal_one_update_id,
            PublicationRecord.linked_lawsuit_id,
            PublicationRecord.publication_date,
        ).filter(PublicationRecord.legal_one_update_id < 0).all()
    }
    no_lote: dict[int, tuple[Optional[int], Optional[str]]] = {}

    for linha in linhas:
        chave = (linha["lawsuit_id"], linha["publication_date"])
        base = _hash_base(
            linha["lawsuit_id"], linha["publication_date"], linha["descricao"]
        )
        passo = 0
        while True:
            candidato = _id_sintetico(base, passo)
            dono = ocupados.get(candidato, no_lote.get(candidato))
            if dono is None or dono == chave:
                break
            passo += 1
            if passo > 1000:  # inatingível na prática; evita laço infinito
                logger.error(
                    "Colisão persistente de ID sintético para o processo %s.",
                    linha["lawsuit_id"],
                )
                break
        linha["update_id"] = candidato
        no_lote[candidato] = chave


# ── Leitura da planilha ────────────────────────────────────────────────

def ler_planilha(conteudo: bytes, db: Session) -> dict[str, Any]:
    """Lê o .xlsx e devolve o resultado da análise, SEM gravar nada.

    Serve tanto para a pré-visualização quanto para a importação — a
    importação usa exatamente o mesmo resultado, então o que o operador vê no
    preview é literalmente o que vai entrar.
    """
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - dependência é fixa
        raise ValueError(f"Leitor de planilha indisponível: {exc}")

    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(conteudo), read_only=True, data_only=True
        )
    except Exception as exc:
        raise ValueError(
            "Não foi possível abrir o arquivo. Confirme que é um .xlsx "
            f"exportado do Legal One. Detalhe: {exc}"
        )

    ws = wb[wb.sheetnames[0]]
    iterador = ws.iter_rows(values_only=True)
    try:
        header = next(iterador)
    except StopIteration:
        raise ValueError("A planilha está vazia.")

    col, faltando = _mapear_colunas(header)
    if faltando:
        rotulos = {
            "lawsuit_id": "Id",
            "cnj": "Nº do processo",
            "data": "Andamentos / Data/hora",
            "descricao": "Andamentos / Descrição",
        }
        nomes = ", ".join(rotulos.get(c, c) for c in faltando)
        raise ValueError(
            f"A planilha não tem a(s) coluna(s) obrigatória(s): {nomes}. "
            "Confira se a extração do Legal One inclui a coluna 'Id' "
            "(identificador do processo)."
        )

    # path → external_id. O `path` é a hierarquia completa e é a convenção da
    # casa pra identificar escritório (o `name` é só a folha e se repete).
    escritorios = {
        _norm(o.path): o.external_id
        for o in db.query(LegalOneOffice).filter(LegalOneOffice.path.isnot(None)).all()
    }

    def celula(linha: tuple, campo: str) -> Any:
        idx = col.get(campo)
        if idx is None or idx >= len(linha):
            return None
        return linha[idx]

    validas: list[dict] = []
    ignoradas: list[dict] = []
    escritorios_sem_match: set[str] = set()
    vistas_no_arquivo: set[tuple] = set()
    total_linhas = 0

    for numero, linha in enumerate(iterador, start=2):
        if linha is None or not any(
            c is not None and str(c).strip() != "" for c in linha
        ):
            continue
        total_linhas += 1
        if total_linhas > MAX_LINHAS:
            raise ValueError(
                f"A planilha tem mais de {MAX_LINHAS:,} linhas. "
                "Divida a extração em períodos menores.".replace(",", ".")
            )

        # Só andamentos do tipo Publicação (quando a coluna existe).
        tipo = celula(linha, "tipo")
        if "tipo" in col and tipo is not None and str(tipo).strip() != "":
            if _norm(tipo) != TIPO_PUBLICACAO:
                ignoradas.append(
                    {"linha": numero, "motivo": f"tipo de andamento '{tipo}'"}
                )
                continue

        bruto_id = celula(linha, "lawsuit_id")
        try:
            lawsuit_id = int(str(bruto_id).strip().split(".")[0])
            if lawsuit_id <= 0:
                raise ValueError
        except (TypeError, ValueError):
            ignoradas.append(
                {"linha": numero, "motivo": "coluna Id vazia ou não numérica"}
            )
            continue

        publication_date = _data_publicacao(celula(linha, "data"))
        if not publication_date:
            ignoradas.append(
                {"linha": numero, "motivo": "data do andamento vazia ou inválida"}
            )
            continue

        descricao = str(celula(linha, "descricao") or "").strip()
        if not descricao:
            ignoradas.append({"linha": numero, "motivo": "descrição vazia"})
            continue

        cnj_digitos = _so_digitos(celula(linha, "cnj"))
        cnj = str(celula(linha, "cnj") or "").strip() or None
        if len(cnj_digitos) != 20:
            # Não é motivo pra descartar: o processo já está identificado pelo
            # Id, que é o que importa. O CNJ é enriquecido depois pelo L1.
            logger.debug("Linha %s: CNJ fora do padrão (%r).", numero, cnj)

        path_escritorio = str(celula(linha, "escritorio") or "").strip()
        office_id = escritorios.get(_norm(path_escritorio)) if path_escritorio else None
        if path_escritorio and office_id is None:
            escritorios_sem_match.add(path_escritorio)

        # Linha idêntica repetida dentro do MESMO arquivo (o operador às vezes
        # cola duas extrações na mesma aba) — conta uma vez só.
        assinatura = (lawsuit_id, publication_date, hashlib.sha1(
            descricao.encode("utf-8")).hexdigest())
        if assinatura in vistas_no_arquivo:
            ignoradas.append(
                {"linha": numero, "motivo": "linha repetida dentro do arquivo"}
            )
            continue
        vistas_no_arquivo.add(assinatura)

        validas.append({
            "linha": numero,
            "lawsuit_id": lawsuit_id,
            "cnj": cnj,
            "pasta": str(celula(linha, "pasta") or "").strip() or None,
            "escritorio_path": path_escritorio or None,
            "office_id": office_id,
            "responsavel": str(celula(linha, "responsavel") or "").strip() or None,
            "publication_date": publication_date,
            "descricao": descricao,
            "data_cadastro": _data_iso(celula(linha, "data_cadastro")),
        })

    wb.close()

    if not validas:
        raise ValueError(
            "Nenhuma linha aproveitável na planilha. "
            f"{len(ignoradas)} linha(s) ignorada(s)."
        )

    atribuir_update_ids(db, validas)

    datas = sorted({v["publication_date"][:10] for v in validas})
    por_escritorio: dict[str, int] = {}
    for v in validas:
        rotulo = v["escritorio_path"] or "(sem escritório)"
        por_escritorio[rotulo] = por_escritorio.get(rotulo, 0) + 1

    return {
        "validas": validas,
        "ignoradas": ignoradas,
        "total_linhas": total_linhas,
        "total_validas": len(validas),
        "total_ignoradas": len(ignoradas),
        "processos_distintos": len({v["lawsuit_id"] for v in validas}),
        "data_inicial": datas[0] if datas else None,
        "data_final": datas[-1] if datas else None,
        "por_escritorio": dict(
            sorted(por_escritorio.items(), key=lambda kv: -kv[1])
        ),
        "escritorios_nao_encontrados": sorted(escritorios_sem_match),
    }


def montar_publicacoes(validas: list[dict]) -> list[dict]:
    """Converte as linhas válidas para o contrato de publicação do Legal One.

    A partir daqui o resto do sistema não sabe (nem precisa saber) que a origem
    foi uma planilha: o dicionário tem o mesmo formato do que vem da API.
    """
    publicacoes: list[dict] = []
    for v in validas:
        publicacoes.append({
            "id": v["update_id"],
            "originType": ORIGIN_TYPE,
            "typeId": UPDATE_TYPE_ID,
            "description": v["descricao"],
            "notes": None,
            "date": v["publication_date"],
            "creationDate": v["publication_date"],
            "relationships": [
                {"linkType": "Litigation", "linkId": v["lawsuit_id"]}
            ],
            # Chaves privadas do pipeline. Se o L1 estiver de pé, o
            # `_enrich_with_lawsuit_data` sobrescreve com o dado oficial; se
            # estiver fora, o enriquecimento falha soft e esses valores da
            # planilha permanecem. Funciona nos dois cenários.
            "_cnj": v["cnj"],
            "_responsible_office_id": v["office_id"],
            "_lawsuit_creation_date": v["data_cadastro"],
        })
    return publicacoes

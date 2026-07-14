"""Acessório "Análise de Risco" da Tarefas por Planilha.

O banco (BB) manda a planilha de agendamento no formato-modelo, mas SEM o CNJ
(só o NPJ). A "Base Analítica" traz o relacional NPJ -> Nº do Processo (CNJ) +
a coluna "Validação nº CNJ" (CNJ CORRETO/INCORRETO). Este módulo faz o join por
NPJ e devolve a MESMA planilha de agendamento com a coluna CNJ preenchida, pronta
pra subir em Tarefas por Planilha.

Regra de resolução por NPJ (um NPJ pode ter >1 processo na base — coluna
"Variação"):
  - 1 CNJ marcado "CNJ CORRETO"            -> preenche (ok)
  - vários "CNJ CORRETO" distintos          -> deixa em branco (ambíguo, revisar)
  - nenhum correto, 1 número só             -> em branco + revisar (cnj_incorreto)
  - nenhum correto, vários números          -> em branco (ambíguo, revisar)
  - NPJ ausente na base                     -> em branco (não encontrado, revisar)
  - linha da análise que JÁ tinha CNJ       -> preserva o que veio (ja_tinha_cnj)

Preserva a planilha original (formatação/colunas) e acrescenta as abas "Revisar"
e "Resumo". Não toca no L1 — é um transform de planilha puro.
"""
from __future__ import annotations

from io import BytesIO
from typing import Dict, List, Tuple

from openpyxl import Workbook, load_workbook

VALIDACAO_OK = "cnj correto"


def _norm(v) -> str:
    """Normaliza célula pra chave/comparação: str, sem espaços, sem `.0` de float."""
    if v is None:
        return ""
    s = str(v).strip()
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s


def _col_index(headers: List[str], *nomes: str) -> int | None:
    """Acha a coluna pelo nome do cabeçalho (case-insensitive, primeira que casar)."""
    low = [(_norm(h) or "").lower() for h in headers]
    for nome in nomes:
        alvo = nome.lower()
        for i, h in enumerate(low):
            if h == alvo:
                return i
    # fallback: startswith (ex.: "Nº do Processo")
    for nome in nomes:
        alvo = nome.lower()
        for i, h in enumerate(low):
            if h.startswith(alvo):
                return i
    return None


def _carregar_base(base_bytes: bytes) -> Dict[str, List[Tuple[str, str]]]:
    """NPJ -> lista de (nº do processo/CNJ, validação). read_only pela base ser grande."""
    wb = load_workbook(BytesIO(base_bytes), read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    try:
        header = list(next(it))
    except StopIteration:
        wb.close()
        raise ValueError("Base Analítica vazia (sem cabeçalho).")
    i_npj = _col_index(header, "NPJ")
    i_proc = _col_index(header, "Nº do Processo", "N° do Processo", "Numero do Processo", "CNJ")
    i_val = _col_index(header, "Validação nº CNJ", "Validacao nº CNJ", "Validação")
    if i_npj is None or i_proc is None:
        wb.close()
        raise ValueError(
            "A Base Analítica precisa das colunas 'NPJ' e 'Nº do Processo'. "
            f"Cabeçalho lido: {[_norm(h) for h in header][:8]}…"
        )

    def g(row, idx):
        return row[idx] if idx is not None and idx < len(row) else None

    mapa: Dict[str, List[Tuple[str, str]]] = {}
    for row in it:
        if not row or all(c is None for c in row):
            continue
        npj = _norm(g(row, i_npj))
        if not npj:
            continue
        proc = _norm(g(row, i_proc))
        val = (_norm(g(row, i_val)) or "").lower()
        mapa.setdefault(npj, []).append((proc, val))
    wb.close()
    return mapa


def _resolver(npj: str, mapa: Dict[str, List[Tuple[str, str]]]) -> Tuple[str, str, List[str]]:
    """Devolve (cnj, status, candidatos). status ∈ ok|ambiguo|cnj_incorreto|nao_encontrado."""
    candidatos = mapa.get(npj)
    if not candidatos:
        return "", "nao_encontrado", []
    corretos = sorted({p for p, v in candidatos if p and v == VALIDACAO_OK})
    todos = sorted({p for p, v in candidatos if p})
    if len(corretos) == 1:
        return corretos[0], "ok", corretos
    if len(corretos) > 1:
        return "", "ambiguo", corretos
    if len(todos) == 1:
        return "", "cnj_incorreto", todos  # existe 1 número, mas reprovou validação
    if len(todos) > 1:
        return "", "ambiguo", todos
    return "", "nao_encontrado", []


_MOTIVOS = {
    "ambiguo": "Mais de um CNJ para o NPJ — escolher manualmente",
    "cnj_incorreto": "Único CNJ da base reprovou na validação (CNJ INCORRETO)",
    "nao_encontrado": "NPJ não encontrado na Base Analítica",
}


def gerar_planilha_com_cnj(analise_bytes: bytes, base_bytes: bytes) -> Tuple[bytes, dict]:
    """Faz o join NPJ->CNJ e devolve (xlsx_bytes, resumo)."""
    mapa = _carregar_base(base_bytes)

    wb = load_workbook(BytesIO(analise_bytes))
    ws = wb.active
    header = [c.value for c in ws[1]]
    i_npj = _col_index(header, "NPJ")
    i_cnj = _col_index(header, "CNJ")
    if i_npj is None or i_cnj is None:
        raise ValueError(
            "A planilha de Análise de Risco precisa das colunas 'NPJ' e 'CNJ' "
            "(a CNJ vem em branco e é preenchida por aqui)."
        )
    col_npj, col_cnj = i_npj + 1, i_cnj + 1  # openpyxl é 1-based

    resumo = {"total": 0, "resolvidos": 0, "ja_tinha_cnj": 0,
              "ambiguo": 0, "cnj_incorreto": 0, "nao_encontrado": 0}
    revisar: List[dict] = []

    for r in range(2, ws.max_row + 1):
        npj = _norm(ws.cell(row=r, column=col_npj).value)
        cnj_atual = _norm(ws.cell(row=r, column=col_cnj).value)
        # linha vazia de verdade (sem NPJ e sem CNJ) — ignora
        if not npj and not cnj_atual:
            continue
        resumo["total"] += 1
        if cnj_atual:
            resumo["ja_tinha_cnj"] += 1
            continue
        cnj, status, candidatos = _resolver(npj, mapa)
        if status == "ok":
            ws.cell(row=r, column=col_cnj, value=cnj)
            resumo["resolvidos"] += 1
        else:
            resumo[status] += 1
            revisar.append({
                "linha": r, "npj": npj,
                "motivo": _MOTIVOS.get(status, status),
                "candidatos": " || ".join(candidatos) if candidatos else "—",
            })

    # Aba Revisar
    if revisar:
        wr = wb.create_sheet("Revisar")
        wr.append(["Linha", "NPJ", "Motivo", "CNJ candidatos"])
        for it in revisar:
            wr.append([it["linha"], it["npj"], it["motivo"], it["candidatos"]])
        for col, w in zip("ABCD", (8, 16, 46, 60)):
            wr.column_dimensions[col].width = w

    # Aba Resumo
    wr = wb.create_sheet("Resumo")
    wr.append(["Métrica", "Valor"])
    linhas_resumo = [
        ("Total de linhas", resumo["total"]),
        ("CNJ preenchido (resolvido)", resumo["resolvidos"]),
        ("Já tinham CNJ", resumo["ja_tinha_cnj"]),
        ("Revisar — ambíguo (vários CNJ)", resumo["ambiguo"]),
        ("Revisar — CNJ incorreto", resumo["cnj_incorreto"]),
        ("Revisar — NPJ não encontrado", resumo["nao_encontrado"]),
    ]
    for k, v in linhas_resumo:
        wr.append([k, v])
    wr.column_dimensions["A"].width = 34
    wr.column_dimensions["B"].width = 12

    out = BytesIO()
    wb.save(out)
    wb.close()
    resumo["revisar_total"] = len(revisar)
    return out.getvalue(), resumo

"""Carga do ESTOQUE da Equipe Mista a partir da Base Analítica do BB.

Diferente do fluxo diário (que descobre vínculo no ato da captura), aqui a gente
importa de uma vez o que já está em curso: parte (CPF/CNPJ) com processos nossos
nos DOIS polos do BB. Cada parte vira um grupo, o grupo inteiro vai pra UMA
advogada (nunca se divide um cliente entre duas), e as pastas entram como
vínculo com `transicao_pendente` — o painel dimensiona, o supervisor conduz o
handover. **Nada é redistribuído no Legal One por aqui.**

Formato dos polos na Analítica: "NOME | MCI | CPF-ou-CNPJ", partes separadas por
"||". O documento vem SEM zeros à esquerda (o BB aparece como "191").
"""
from __future__ import annotations

import logging
import re
import unicodedata
from collections import defaultdict
from datetime import datetime, timezone
from typing import Any, Optional

from sqlalchemy.orm import Session

logger = logging.getLogger("distribuidos_bb.estoque_mista")

# Partes institucionais: aparecem nos dois polos por litisconsórcio, não porque
# são "o mesmo devedor dos dois lados". Casa por substring no nome, sem acento.
TERMOS_INSTITUCIONAIS = [
    "banco ", "banco do", "bradesco", "itau", "santander", "caixa economica",
    "bancoob", "banrisul", "safra", "daycoval", "banco pan", "bmg",
    "votorantim", "sicredi", "sicoob", "nubank", "banco da amazonia",
    "seguros", "seguradora", "capitalizacao", "previdencia", "resseguros",
    "consorcio", "consorcios", "leasing", "arrendamento mercantil",
    "administradora de cartoes", "cartoes de credito", "financeira",
    "credito imobiliario", "distribuidora de titulos", "corretora",
    "fundo de investimento", "fundo de arrendamento", "securitizadora",
    "cia de seguros", "companhia de seguros", "alianca do brasil",
    "brasilprev", "brasilcap", "brasilveiculos", "mapfre",
    "inss", "instituto nacional", "fazenda nacional", "uniao federal",
    "municipio", "estado do", "governo do estado", "prefeitura",
    "ministerio publico", "defensoria", "procuradoria",
    "fundo nacional", "desenvolvimento da educacao", "bndes",
]


def _dig(v: Any) -> str:
    return re.sub(r"\D", "", str(v or ""))


def _sem_acento(v: Any) -> str:
    s = unicodedata.normalize("NFKD", str(v or "").lower())
    return "".join(c for c in s if not unicodedata.combining(c))


def eh_institucional(nome: Optional[str]) -> bool:
    n = _sem_acento(nome)
    return any(t in n for t in TERMOS_INSTITUCIONAIS)


def normalizar_doc(doc: Any) -> Optional[str]:
    """Zero-pad pro tamanho canônico: 11 = CPF, 14 = CNPJ. None se inválido."""
    d = _dig(doc)
    if not d:
        return None
    if len(d) <= 11:
        d = d.zfill(11)
    elif len(d) <= 14:
        d = d.zfill(14)
    else:
        return None
    if len(set(d)) == 1:  # 000..., 111... = placeholder da base
        return None
    return d


def parse_polo(valor: Any) -> list[dict[str, Any]]:
    """'NOME | MCI | DOC || NOME2 | MCI2 | DOC2' → [{nome, mci, doc}, ...]"""
    out: list[dict[str, Any]] = []
    if not valor:
        return out
    for bloco in str(valor).split("||"):
        segs = [s.strip() for s in bloco.split("|") if s.strip()]
        if not segs:
            continue
        out.append({
            "nome": segs[0],
            "mci": segs[1] if len(segs) > 1 else None,
            "doc": normalizar_doc(segs[2]) if len(segs) > 2 else None,
        })
    return out


def ler_analitica(conteudo: bytes) -> list[dict[str, Any]]:
    """Lê a Base Analítica (aba Export) e devolve os processos com as partes."""
    import io

    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(conteudo), data_only=True, read_only=True)
    ws = wb["Export"] if "Export" in wb.sheetnames else wb.worksheets[0]
    it = ws.iter_rows(values_only=True)
    hdr = [str(c or "").strip() for c in next(it)]
    ix = {h: i for i, h in enumerate(hdr)}

    def cel(row, nome):
        i = ix.get(nome)
        return row[i] if (i is not None and i < len(row)) else None

    processos: list[dict[str, Any]] = []
    for r in it:
        if not r or not r[0]:
            continue
        partes: list[dict[str, Any]] = []
        for campo in ("Polo Ativo", "Polo Passivo"):
            partes.extend(parse_polo(cel(r, campo)))
        processos.append({
            "npj": _dig(cel(r, "NPJ")),
            # CNJ verbatim: a validação do BB erra, a regra da casa é usar como veio.
            "cnj": _dig(cel(r, "N° do Processo")),
            "polo_bb": str(cel(r, "Polo do BB") or "").upper(),
            "materia": cel(r, "Matéria"),
            "acao": cel(r, "Tipo de Ação"),
            "situacao": cel(r, "Situação do Processo"),
            "natureza": cel(r, "Natureza"),
            "uf": cel(r, "UF"),
            "valor": cel(r, "Valor da Causa"),
            "partes": partes,
        })
    wb.close()
    return processos


def identificar_alvo(processos: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    """Partes com processos nos DOIS polos, já sem as institucionais.

    Devolve {doc: {"nome":…, "procs":[…]}}.
    """
    por_doc: dict[str, dict[str, Any]] = defaultdict(lambda: {"nome": None, "procs": []})
    for p in processos:
        vistos: set[str] = set()
        for parte in p["partes"]:
            doc = parte.get("doc")
            if not doc or doc in vistos or eh_institucional(parte.get("nome")):
                continue
            vistos.add(doc)
            por_doc[doc]["nome"] = por_doc[doc]["nome"] or parte.get("nome")
            por_doc[doc]["procs"].append(p)

    alvo: dict[str, dict[str, Any]] = {}
    for doc, v in por_doc.items():
        if len(v["procs"]) < 2:
            continue
        polos = {p["polo_bb"] for p in v["procs"]}
        if "AUTOR" in polos and "REU" in polos:
            alvo[doc] = v
    return alvo


def agrupar_clusters(alvo: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    """Une partes que COMPARTILHAM processo (cônjuges, sócios, litisconsortes).

    Sem isso, duas partes do mesmo processo cairiam pra advogadas diferentes e a
    pasta ficaria com as duas — pior do que dividir cliente. O cluster conectado
    mantém o grupo familiar/societário inteiro com uma só advogada.

    Devolve [{"docs": [...], "cnjs": {...}, "peso": n}], onde `peso` é a
    quantidade de pastas DISTINTAS do cluster.
    """
    pai: dict[str, str] = {d: d for d in alvo}

    def find(x: str) -> str:
        while pai[x] != x:
            pai[x] = pai[pai[x]]
            x = pai[x]
        return x

    def union(a: str, b: str) -> None:
        ra, rb = find(a), find(b)
        if ra != rb:
            pai[rb] = ra

    # Aresta entre partes que aparecem no mesmo CNJ
    por_cnj: dict[str, list[str]] = defaultdict(list)
    for doc, v in alvo.items():
        for p in v["procs"]:
            if p["cnj"]:
                por_cnj[p["cnj"]].append(doc)
    for docs in por_cnj.values():
        for outro in docs[1:]:
            union(docs[0], outro)

    grupos: dict[str, list[str]] = defaultdict(list)
    for doc in alvo:
        grupos[find(doc)].append(doc)

    clusters: list[dict[str, Any]] = []
    for docs in grupos.values():
        cnjs = {p["cnj"] for d in docs for p in alvo[d]["procs"] if p["cnj"]}
        clusters.append({"docs": docs, "cnjs": cnjs, "peso": len(cnjs)})
    return clusters


def dividir_entre_advogadas(
    alvo: dict[str, dict[str, Any]], advogadas: list[int]
) -> dict[str, int]:
    """LPT sobre os CLUSTERS: maiores primeiro, cada um pro lado mais leve.

    O cluster inteiro (todas as partes ligadas por processo em comum) vai pra
    UMA advogada — nunca se divide um cliente nem se duplica uma pasta.
    Devolve {doc: user_id}.
    """
    if not advogadas:
        return {}
    clusters = sorted(agrupar_clusters(alvo), key=lambda c: -c["peso"])
    carga = {uid: 0 for uid in advogadas}
    atribuicao: dict[str, int] = {}
    for c in clusters:
        escolhida = min(carga, key=lambda uid: (carga[uid], uid))
        for doc in c["docs"]:
            atribuicao[doc] = escolhida
        carga[escolhida] += c["peso"]
    return atribuicao


def carregar_estoque(
    db: Session,
    conteudo: bytes,
    *,
    advogadas_user_ids: list[int],
    dry_run: bool = True,
) -> dict[str, Any]:
    """Importa o estoque: identifica o alvo, casa no L1, divide e grava vínculos.

    `dry_run=True` (default) só dimensiona e devolve o relatório, sem gravar.
    """
    from app.models.distribuidos_bb import (
        CLIENTE_BB,
        PROC_DISTRIBUIDO,
        POOL_CADASTRADO_L1,
        VINCULO_CENARIO_1,
        BbProcesso,
        BbVinculo,
    )
    from app.models.legal_one import LegalOneOffice, LegalOneUser
    from app.services.legal_one_client import LegalOneApiClient

    processos = ler_analitica(conteudo)
    alvo = identificar_alvo(processos)
    atribuicao = dividir_entre_advogadas(alvo, advogadas_user_ids)

    # CNJs distintos (um processo com 2 partes do alvo aparece 2x — dedupe aqui)
    cnjs: set[str] = {p["cnj"] for v in alvo.values() for p in v["procs"] if p["cnj"]}

    # Casa com o L1 pelo cache (CNJ → lawsuit_id + escritório)
    from sqlalchemy import text

    cache: dict[str, tuple[int, Any]] = {}
    for lid, payload in db.execute(text("select lawsuit_id, payload from lawsuit_cache")):
        if isinstance(payload, dict):
            c = _dig(payload.get("identifierNumber"))
            if c:
                cache[c] = (lid, payload.get("responsibleOfficeId"))
    achados = {c: cache[c] for c in cnjs if c in cache}

    offices = {o.external_id: (o.path or o.name) for o in db.query(LegalOneOffice).all()}
    nomes_adv = {u.id: u.name for u in db.query(LegalOneUser)
                 .filter(LegalOneUser.id.in_(advogadas_user_ids or [0])).all()}

    clusters = agrupar_clusters(alvo)
    relatorio: dict[str, Any] = {
        "processos_analitica": len(processos),
        "partes_alvo": len(alvo),
        "clusters": len(clusters),
        "pastas_distintas": len(cnjs),
        "cnjs_no_l1": len(achados),
        "cnjs_fora_do_l1": len(cnjs) - len(achados),
        "por_advogada": {},
        "dry_run": dry_run,
    }
    # Pastas por advogada contadas em CNJs DISTINTOS (uma pasta compartilhada
    # entre partes do mesmo cluster conta uma vez só).
    pastas_por_uid: dict[int, set[str]] = defaultdict(set)
    partes_por_uid: dict[int, int] = defaultdict(int)
    for doc, uid in atribuicao.items():
        partes_por_uid[uid] += 1
        for p in alvo[doc]["procs"]:
            if p["cnj"]:
                pastas_por_uid[uid].add(p["cnj"])
    for uid, docs_set in pastas_por_uid.items():
        relatorio["por_advogada"][nomes_adv.get(uid, str(uid))] = {
            "partes": partes_por_uid[uid],
            "pastas": len(docs_set),
        }
    # Conferência: nenhuma pasta pode estar com duas advogadas.
    todas = [c for s in pastas_por_uid.values() for c in s]
    relatorio["pastas_em_duas_advogadas"] = len(todas) - len(set(todas))

    if dry_run:
        return relatorio

    # ── Gravação ───────────────────────────────────────────────────────────
    # Responsável atual de cada pasta (aquece o cache; usa a API só nos misses).
    client = LegalOneApiClient()
    lawsuit_ids = [lid for lid, _off in achados.values()]
    try:
        responsaveis = client.prefetch_lawsuit_responsibles_cache(lawsuit_ids)
    except Exception:  # noqa: BLE001
        logger.exception("Estoque Mista: falha ao buscar responsáveis no L1.")
        responsaveis = {}

    users_por_nome = {
        _sem_acento(u.name): u.id
        for u in db.query(LegalOneUser).filter(LegalOneUser.name.isnot(None)).all()
    }
    agora = datetime.now(timezone.utc)
    criados_proc = criados_vinc = 0

    for doc, v in alvo.items():
        uid = atribuicao.get(doc)
        # Um "processo guarda-chuva" por PARTE: é o agrupador do painel.
        fp = "estoque:mista:%s" % doc
        proc = db.query(BbProcesso).filter(BbProcesso.fingerprint == fp).first()
        if proc is None:
            proc = BbProcesso(cliente=CLIENTE_BB, fingerprint=fp,
                              status=PROC_DISTRIBUIDO, planilha_status=POOL_CADASTRADO_L1)
            db.add(proc)
            criados_proc += 1
        proc.adverso_principal = v["nome"]
        proc.responsavel_user_id = uid
        proc.vinculo_cenario = VINCULO_CENARIO_1
        proc.vinculos_qtd = len(v["procs"])
        proc.vinculos_verificado_em = agora
        proc.observacao = "Estoque"
        db.flush()

        db.query(BbVinculo).filter(BbVinculo.processo_id == proc.id).delete(
            synchronize_session=False
        )
        for p in v["procs"]:
            lid, off = achados.get(p["cnj"], (None, None))
            resp = responsaveis.get(lid) if lid else None
            resp_nome = (resp or {}).get("name")
            resp_id = users_por_nome.get(_sem_acento(resp_nome)) if resp_nome else None
            db.add(BbVinculo(
                processo_id=proc.id,
                doc_parte=doc,
                nome_parte=v["nome"],
                npj=p["npj"] or "-",
                numero_processo=p["npj"],
                cnj=p["cnj"] or None,
                contrario_nome=v["nome"],
                situacao=p["situacao"],
                natureza=p["natureza"],
                uja=None,
                polo=("Ativo" if p["polo_bb"] == "AUTOR" else "Passivo"),
                posicao_banco=("Autor" if p["polo_bb"] == "AUTOR" else "Réu"),
                l1_lawsuit_id=lid,
                l1_folder=offices.get(off) and None,  # folder real vem do monitor
                responsavel_atual_user_id=resp_id,
                responsavel_atual_nome=resp_nome,
                na_equipe_mista=bool(resp_id and resp_id in (advogadas_user_ids or [])),
                # Estoque inteiro entra pendente: o handover é do supervisor.
                transicao_pendente=True,
                raw={"origem": "estoque_analitica", "materia": p["materia"],
                     "acao": p["acao"], "uf": p["uf"], "valor": p["valor"],
                     "escritorio_l1": offices.get(off)},
            ))
            criados_vinc += 1
        db.commit()

    relatorio["processos_criados"] = criados_proc
    relatorio["vinculos_criados"] = criados_vinc
    relatorio["responsaveis_resolvidos"] = len(responsaveis)
    return relatorio

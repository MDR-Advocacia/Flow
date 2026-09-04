"""Export da base NERC — o retrato da carteira especializada em planilha.

Não confundir com a "Gerar planilha" do topo da página: aquela é a planilha de
MIGRAÇÃO, no formato que o import do Legal One consome pra criar pasta. Esta é
um relatório legível, pra conferência e acompanhamento — a base do NERC como o
operador a enxerga no painel Réu/Autor.

A unidade da planilha é a PASTA, não o processo. Um vínculo é uma relação entre
duas pastas (a que acabou de entrar e a antiga da mesma parte), e as duas
compõem a carteira: listar só o processo novo esconderia metade dela.

Duas abas:
  - "Base NERC"  — uma linha por pasta, com a origem marcada (processo novo ou
    pasta vinculada) e a parte que amarra as duas;
  - "Por parte"  — a mesma base agrupada, que é como a operação pensa a
    carteira: quem é a parte, quantas pastas ela tem e quem conduz.
"""
from __future__ import annotations

import io
from typing import Any, Optional

from sqlalchemy.orm import Session

from app.models.distribuidos_bb import BbProcesso, BbVinculo
from app.models.legal_one import LegalOneUser

CENARIO_LABEL = {
    "CENARIO_1": "Novo na equipe (transição pendente)",
    "CENARIO_2": "Parte já especializada",
}

COLS_BASE = [
    ("Origem", 20), ("Parte", 38), ("Documento da parte", 20),
    ("NPJ", 20), ("CNJ", 26), ("Pasta L1", 14),
    ("Posição (banco)", 15), ("Situação", 16),
    ("Responsável atual", 28), ("Cenário", 32),
    ("Etiqueta NERC", 18), ("Transição", 30), ("Capturado em", 18),
]
COLS_PARTE = [
    ("Parte", 40), ("Documento", 20), ("Pastas na carteira", 18),
    ("Responsável do processo novo", 28), ("Cenário", 32),
    ("Transições pendentes", 20), ("NPJ do processo novo", 20),
]


def _dt(x) -> str:
    return x.strftime("%d/%m/%Y %H:%M") if x else ""


def _cnj(valor: Optional[str]) -> str:
    """CNJ com máscara. O processo novo já vem formatado do portal, o vínculo
    vem cru do Legal One — sem isto a mesma coluna mistura os dois formatos e
    não dá pra procurar um número na planilha."""
    if not valor:
        return ""
    d = "".join(ch for ch in valor if ch.isdigit())
    if len(d) != 20:
        return valor
    return f"{d[:7]}-{d[7:9]}.{d[9:13]}.{d[13:14]}.{d[14:16]}.{d[16:]}"


def _doc(valor: Optional[str]) -> str:
    """CPF/CNPJ com máscara, completando os zeros à esquerda POR TAMANHO.

    O portal do BB entrega o documento sem os zeros da frente (CPF 097.043.702-10
    chega como "9704370210"), e foi exatamente isso que descartou 21% das partes
    na pesquisa de vínculos. Aqui o efeito é só cosmético — mas um CPF com 10
    dígitos na planilha não casa com nada que o operador for conferir.
    """
    if not valor:
        return ""
    d = "".join(ch for ch in valor if ch.isdigit())
    if not d:
        return valor
    if len(d) <= 11:
        d = d.zfill(11)
        return f"{d[:3]}.{d[3:6]}.{d[6:9]}-{d[9:]}"
    d = d.zfill(14)
    return f"{d[:2]}.{d[2:5]}.{d[5:8]}/{d[8:12]}-{d[12:]}"


def _parte_do_processo(vinculos: list[BbVinculo], proc: BbProcesso) -> tuple[str, str]:
    """Nome e documento da parte que gerou o vínculo.

    Vem do vínculo, não do processo: `adverso_principal` é texto livre do portal
    do BB e nem sempre casa com a parte pesquisada — o vínculo guarda o nome e o
    CPF/CNPJ que realmente casaram no Legal One.
    """
    for v in vinculos:
        if v.nome_parte or v.doc_parte:
            return (v.nome_parte or ""), (v.doc_parte or "")
    return (proc.adverso_principal or ""), ""


def montar_xlsx_base_nerc(
    db: Session,
    rows: list[BbProcesso],
    vincs: dict[int, list[BbVinculo]],
) -> tuple[io.BytesIO, int]:
    """Monta a planilha. Devolve (buffer, total de linhas da aba principal)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    resp_ids = {p.responsavel_user_id for p in rows if p.responsavel_user_id}
    resp_ids |= {
        v.responsavel_atual_user_id
        for lista in vincs.values() for v in lista
        if v.responsavel_atual_user_id
    }
    nomes = dict(
        db.query(LegalOneUser.id, LegalOneUser.name)
        .filter(LegalOneUser.id.in_(resp_ids or {0}))
        .all()
    )

    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    # O processo novo é a âncora da parte; destacá-lo evita ler a planilha como
    # se todas as linhas fossem entradas independentes.
    fill_novo = PatternFill("solid", fgColor="FFF4CE")

    def _cabecalho(ws, cols):
        for ci, (nome, w) in enumerate(cols, start=1):
            c = ws.cell(row=1, column=ci, value=nome)
            c.fill, c.font = header_fill, header_font
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[1].height = 22

    # ── aba 1: uma linha por pasta ────────────────────────────────────────
    ws = wb.active
    ws.title = "Base NERC"
    _cabecalho(ws, COLS_BASE)

    ri = 1
    for p in rows:
        lista = vincs.get(p.id, [])
        parte_nome, parte_doc = _parte_do_processo(lista, p)
        cenario = CENARIO_LABEL.get(p.vinculo_cenario or "", p.vinculo_cenario or "")

        ri += 1
        linhas: list[list[Any]] = [[
            "Processo novo", parte_nome, _doc(parte_doc), p.npj or "", _cnj(p.cnj),
            p.l1_folder or "", p.posicao or "", "",
            nomes.get(p.responsavel_user_id) or "", cenario,
            _dt(p.nerc_etiquetado_em) or "não etiquetada", "—", _dt(p.created_at),
        ]]
        for v in lista:
            if v.transicao_pendente:
                transicao = "PENDENTE — aguardando transferência"
            elif v.transicao_concluida_em:
                destino = nomes.get(v.transicao_para_user_id) if v.transicao_para_user_id else None
                transicao = (f"Transferida para {destino}" if destino else "Concluída") \
                    + f" em {_dt(v.transicao_concluida_em)}"
            else:
                transicao = "—"
            linhas.append([
                "Pasta vinculada", v.nome_parte or parte_nome, _doc(v.doc_parte or parte_doc),
                v.npj or "", _cnj(v.cnj), v.l1_folder or "",
                f"BB {v.posicao_banco}" if v.posicao_banco else "",
                v.situacao or "",
                v.responsavel_atual_nome or (nomes.get(v.responsavel_atual_user_id) or ""),
                cenario,
                _dt(v.nerc_etiquetado_em) or "não etiquetada", transicao, "",
            ])

        for li, vals in enumerate(linhas):
            for ci, val in enumerate(vals, start=1):
                cell = ws.cell(row=ri + li, column=ci, value=val)
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=ci in (2, 10, 12))
                if li == 0:
                    cell.fill = fill_novo
        ri += len(linhas) - 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS_BASE))}{max(1, ri)}"

    # ── aba 2: a carteira por parte ───────────────────────────────────────
    ws2 = wb.create_sheet("Por parte")
    _cabecalho(ws2, COLS_PARTE)
    for i, p in enumerate(rows, start=2):
        lista = vincs.get(p.id, [])
        parte_nome, parte_doc = _parte_do_processo(lista, p)
        vals = [
            parte_nome, _doc(parte_doc),
            len(lista) + 1,  # as pastas antigas + a que acabou de entrar
            nomes.get(p.responsavel_user_id) or "",
            CENARIO_LABEL.get(p.vinculo_cenario or "", p.vinculo_cenario or ""),
            sum(1 for v in lista if v.transicao_pendente),
            p.npj or "",
        ]
        for ci, val in enumerate(vals, start=1):
            cell = ws2.cell(row=i, column=ci, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=ci in (1, 5))
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(COLS_PARTE))}{max(1, len(rows) + 1)}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, ri - 1


def exportar_base_nerc(
    db: Session,
    *,
    cenario: Optional[str] = None,
    transicao: Optional[str] = None,
    busca: Optional[str] = None,
    teto: int = 5000,
) -> tuple[io.BytesIO, int, int]:
    """Aplica os MESMOS filtros do painel e monta a planilha.

    Os filtros são repetidos aqui de propósito: exportar "o que está na tela" só
    é verdade se o recorte for o mesmo. Devolve (buffer, pastas, processos).
    """
    base = db.query(BbProcesso).filter(BbProcesso.vinculo_cenario.isnot(None))
    if cenario:
        base = base.filter(BbProcesso.vinculo_cenario == cenario)
    if transicao == "pendente":
        base = base.filter(
            db.query(BbVinculo.id)
            .filter(BbVinculo.processo_id == BbProcesso.id, BbVinculo.transicao_pendente.is_(True))
            .exists()
        )
    if busca:
        like = f"%{busca.strip()}%"
        base = base.filter(
            BbProcesso.cnj.ilike(like)
            | BbProcesso.npj.ilike(like)
            | BbProcesso.adverso_principal.ilike(like)
        )
    # Sem paginação (é um export), mas com teto: a carteira tem dezenas hoje e
    # cresce por coleta — planilha aberta é melhor que request estourando.
    rows = base.order_by(BbProcesso.id.desc()).limit(teto).all()

    vincs: dict[int, list[BbVinculo]] = {}
    ids = [p.id for p in rows]
    if ids:
        for v in (
            db.query(BbVinculo)
            .filter(BbVinculo.processo_id.in_(ids))
            .order_by(BbVinculo.id)
            .all()
        ):
            vincs.setdefault(v.processo_id, []).append(v)

    buf, pastas = montar_xlsx_base_nerc(db, rows, vincs)
    return buf, pastas, len(rows)

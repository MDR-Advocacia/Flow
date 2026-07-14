"""Consultas e orquestração do módulo Distribuídos BB (usado pelos endpoints).

- Dashboard (KPIs por status/escritório).
- Listagem paginada de processos com filtros.
- Auditoria de um processo (linha do tempo de eventos + envolvidos).
- Log global de eventos (paginado, filtrável por seção/nível).
- Ingestão do output do RPA legado (ponte pra ter dados reais já, antes do
  runner na nuvem): cria/atualiza processos por fingerprint e distribui.
"""
from __future__ import annotations

from typing import Any, Optional

from sqlalchemy import func
from sqlalchemy.orm import Session

from app.models.distribuidos_bb import (
    BbEnvolvido,
    BbEscritorio,
    BbEvento,
    BbPlanilha,
    BbProcesso,
    BbRun,
    NIVEL_SUCESSO,
    POOL_CADASTRADO_L1,
    POOL_NOVO,
    POOL_PENDENTE_CADASTRO,
    PROC_COLETADO,
    PROC_DISTRIBUIDO,
    SECAO_EXTRACAO,
)
from app.models.legal_one import LegalOneUser
from app.services.distribuidos_bb import normalizacao as norm
from app.services.distribuidos_bb.distribuicao_service import distribuir_processo
from app.services.distribuidos_bb.log_service import registrar_evento


class DistribuidosBBService:
    def __init__(self, db: Session):
        self.db = db

    # ── nomes de responsáveis (pra montar DTOs sem N+1) ──────────────────
    def _mapa_nomes(self, user_ids: set[int]) -> dict[int, str]:
        ids = {i for i in user_ids if i}
        if not ids:
            return {}
        rows = (
            self.db.query(LegalOneUser.id, LegalOneUser.name)
            .filter(LegalOneUser.id.in_(ids))
            .all()
        )
        return {rid: nome for rid, nome in rows}

    # ── Dashboard ────────────────────────────────────────────────────────
    def dashboard(self) -> dict[str, Any]:
        por_status = dict(
            self.db.query(BbProcesso.status, func.count(BbProcesso.id))
            .group_by(BbProcesso.status)
            .all()
        )
        total = sum(por_status.values())

        por_escritorio = [
            {"escritorio": path or nome or "—", "total": qtd}
            for path, nome, qtd in (
                self.db.query(
                    BbEscritorio.escritorio_path,
                    BbEscritorio.nome,
                    func.count(BbProcesso.id),
                )
                .join(BbProcesso, BbProcesso.escritorio_id == BbEscritorio.id)
                .group_by(BbEscritorio.escritorio_path, BbEscritorio.nome)
                .all()
            )
        ]

        sem_responsavel = (
            self.db.query(func.count(BbProcesso.id))
            .filter(BbProcesso.responsavel_user_id.is_(None))
            .scalar()
        )
        envolvidos_pendentes = (
            self.db.query(func.count(BbEnvolvido.id))
            .filter(BbEnvolvido.status_contato == "NAO_RESOLVIDO")
            .scalar()
        )

        kpis = {
            "total": total,
            "coletados": por_status.get("COLETADO", 0),
            "ciencia_dada": por_status.get("CIENCIA_DADA", 0),
            "distribuidos": por_status.get("DISTRIBUIDO", 0),
            "cadastrados": por_status.get("CADASTRADO", 0),
            "erros": por_status.get("ERRO", 0),
            "revisao": por_status.get("REVISAO", 0),
            "sem_responsavel": int(sem_responsavel or 0),
            "envolvidos_pendentes": int(envolvidos_pendentes or 0),
        }

        ultima_run = self.db.query(BbRun).order_by(BbRun.id.desc()).first()

        # Planilhas ainda não subidas no Legal One (fica no dashboard inicial).
        planilhas_total = self.db.query(func.count(BbPlanilha.id)).scalar() or 0
        pend_rows = (
            self.db.query(BbPlanilha)
            .filter(BbPlanilha.subido_legalone.is_(False))
            .order_by(BbPlanilha.id.desc())
            .limit(5)
            .all()
        )
        planilhas_pendentes = (
            self.db.query(func.count(BbPlanilha.id))
            .filter(BbPlanilha.subido_legalone.is_(False))
            .scalar()
            or 0
        )
        pool = dict(
            self.db.query(BbProcesso.planilha_status, func.count(BbProcesso.id))
            .filter(BbProcesso.status == PROC_DISTRIBUIDO)
            .group_by(BbProcesso.planilha_status)
            .all()
        )
        planilhas = {
            "total": int(planilhas_total),
            "pendentes": int(planilhas_pendentes),
            "pool_novos": int(pool.get(POOL_NOVO, 0)),
            "pendente_cadastro": int(pool.get(POOL_PENDENTE_CADASTRO, 0)),
            "cadastrado_l1": int(pool.get(POOL_CADASTRADO_L1, 0)),
            "recentes_pendentes": [
                {
                    "id": p.id,
                    "nome_arquivo": p.nome_arquivo,
                    "total_processos": p.total_processos,
                    "origem": p.origem,
                    "created_at": p.created_at.isoformat() if p.created_at else None,
                }
                for p in pend_rows
            ],
        }

        # ── Gráficos do dashboard ────────────────────────────────────────
        por_cliente = [
            {"cliente": cli or "BB", "total": int(q)}
            for cli, q in (
                self.db.query(BbProcesso.cliente, func.count(BbProcesso.id))
                .group_by(BbProcesso.cliente)
                .all()
            )
        ]
        por_natureza = [
            {"natureza": nat or "—", "total": int(q)}
            for nat, q in (
                self.db.query(BbProcesso.natureza, func.count(BbProcesso.id))
                .group_by(BbProcesso.natureza)
                .all()
            )
        ]
        por_posicao = [
            {"posicao": pos or "—", "total": int(q)}
            for pos, q in (
                self.db.query(BbProcesso.posicao, func.count(BbProcesso.id))
                .group_by(BbProcesso.posicao)
                .all()
            )
        ]
        resp_rows = (
            self.db.query(BbProcesso.responsavel_user_id, func.count(BbProcesso.id))
            .group_by(BbProcesso.responsavel_user_id)
            .all()
        )
        nomes_resp = self._mapa_nomes({rid for rid, _ in resp_rows if rid})
        por_responsavel = sorted(
            [
                {"responsavel": nomes_resp.get(rid) or "— sem responsável", "total": int(q)}
                for rid, q in resp_rows
            ],
            key=lambda x: -x["total"],
        )[:12]
        # Estado (UF) — parse da tramitação BB.
        from collections import Counter

        from app.services.distribuidos_bb.cadastro_l1 import parse_tramitacao

        uf_counter: Counter = Counter()
        for (tram,) in self.db.query(BbProcesso.tramitacao).all():
            uf = (parse_tramitacao(tram) or {}).get("uf") or "—"
            uf_counter[uf] += 1
        por_estado = sorted(
            [{"uf": uf, "total": t} for uf, t in uf_counter.items()],
            key=lambda x: -x["total"],
        )
        # Distribuição por data de captura (created_at) — timeline.
        por_data = [
            {"data": str(d), "total": int(q)}
            for d, q in (
                self.db.query(func.date(BbProcesso.created_at), func.count(BbProcesso.id))
                .group_by(func.date(BbProcesso.created_at))
                .order_by(func.date(BbProcesso.created_at))
                .all()
            )
        ]
        ultima_passagem = None
        if ultima_run is not None:
            ultima_passagem = {
                "data": (
                    ultima_run.concluido_em or ultima_run.iniciado_em
                ).isoformat()
                if (ultima_run.concluido_em or ultima_run.iniciado_em)
                else None,
                "capturados": ultima_run.total_coletados,
                "status": ultima_run.status,
            }

        return {
            "kpis": kpis,
            "por_status": por_status,
            "por_escritorio": por_escritorio,
            "ultima_run": self._run_dto(ultima_run) if ultima_run else None,
            "planilhas": planilhas,
            "por_natureza": por_natureza,
            "por_posicao": por_posicao,
            "por_responsavel": por_responsavel,
            "por_estado": por_estado,
            "por_data": por_data,
            "ultima_passagem": ultima_passagem,
            "por_cliente": por_cliente,
        }

    def _run_dto(self, run: BbRun) -> dict[str, Any]:
        return {
            "id": run.id,
            "data_inicial": run.data_inicial,
            "data_final": run.data_final,
            "status": run.status,
            "confirmar_ciencia": run.confirmar_ciencia,
            "total_coletados": run.total_coletados,
            "total_ciencia": run.total_ciencia,
            "total_distribuidos": run.total_distribuidos,
            "total_cadastrados": run.total_cadastrados,
            "total_erros": run.total_erros,
            "iniciado_em": run.iniciado_em.isoformat() if run.iniciado_em else None,
            "concluido_em": run.concluido_em.isoformat() if run.concluido_em else None,
        }

    # ── Listagem de processos ────────────────────────────────────────────
    def _q_processos_ordenada(
        self,
        *,
        status: Optional[str] = None,
        escritorio_id: Optional[int] = None,
        busca: Optional[str] = None,
        planilha_status: Optional[str] = None,
        posicao: Optional[str] = None,
        cliente: Optional[str] = None,
        cadastro_de: Optional[str] = None,
        cadastro_ate: Optional[str] = None,
    ):
        """Query filtrada + ordenada (pendente no topo, depois por data). Reusada
        pela listagem paginada e pela exportação."""
        from datetime import datetime, time as _time

        from sqlalchemy import case

        from app.models.distribuidos_bb import (
            POOL_CADASTRADO_L1,
            POOL_NOVO,
            POOL_PENDENTE_CADASTRO,
        )

        q = self.db.query(BbProcesso)
        if status:
            q = q.filter(BbProcesso.status == status)
        if cliente:
            q = q.filter(BbProcesso.cliente == cliente)
        if planilha_status:
            q = q.filter(BbProcesso.planilha_status == planilha_status)
        if posicao:
            q = q.filter(BbProcesso.posicao == posicao)
        if escritorio_id:
            q = q.filter(BbProcesso.escritorio_id == escritorio_id)
        if busca:
            termo = f"%{busca.strip()}%"
            q = q.filter(
                (BbProcesso.cnj.ilike(termo))
                | (BbProcesso.npj.ilike(termo))
                | (BbProcesso.adverso_principal.ilike(termo))
            )

        def _parse(d: str, fim: bool = False):
            try:
                dt = datetime.strptime(d.strip(), "%Y-%m-%d")
                return datetime.combine(dt.date(), _time.max if fim else _time.min)
            except Exception:  # noqa: BLE001
                return None

        if cadastro_de and (dd := _parse(cadastro_de)):
            q = q.filter(BbProcesso.cadastro_confirmado_em >= dd)
        if cadastro_ate and (da := _parse(cadastro_ate, fim=True)):
            q = q.filter(BbProcesso.cadastro_confirmado_em <= da)

        ordem_pool = case(
            (BbProcesso.planilha_status == POOL_PENDENTE_CADASTRO, 0),
            (BbProcesso.planilha_status == POOL_NOVO, 1),
            (BbProcesso.planilha_status == POOL_CADASTRADO_L1, 2),
            else_=3,
        )
        return q.order_by(
            ordem_pool,
            BbProcesso.cadastro_confirmado_em.desc().nullslast(),
            BbProcesso.id.desc(),
        )

    def listar_processos(self, *, limit: int = 50, offset: int = 0, **filtros) -> dict[str, Any]:
        q = self._q_processos_ordenada(**filtros)
        total = q.count()
        rows = q.limit(limit).offset(offset).all()
        nomes = self._mapa_nomes({r.responsavel_user_id for r in rows})
        return {"total": total, "items": [self._proc_dto(r, nomes) for r in rows]}

    def exportar_processos(self, **filtros):
        """xlsx bonitinho dos processos do filtro atual (todas as infos). Sem
        paginação. Devolve (BytesIO, total)."""
        rows = self._q_processos_ordenada(**filtros).all()
        nomes = self._mapa_nomes({r.responsavel_user_id for r in rows})
        return _montar_xlsx_processos(rows, nomes), len(rows)

    def detalhe_planilha(self, planilha_id: int) -> Optional[dict[str, Any]]:
        """Planilha + seus processos com o status de cadastro no L1 (tela de visualização)."""
        pl = self.db.get(BbPlanilha, planilha_id)
        if pl is None:
            return None
        rows = (
            self.db.query(BbProcesso)
            .filter(BbProcesso.planilha_id == planilha_id)
            .order_by(BbProcesso.id)
            .all()
        )
        nomes = self._mapa_nomes({r.responsavel_user_id for r in rows})
        cadastrados = sum(1 for p in rows if p.planilha_status == POOL_CADASTRADO_L1)
        pendentes = sum(1 for p in rows if p.planilha_status == POOL_PENDENTE_CADASTRO)
        return {
            "planilha": {
                "id": pl.id,
                "nome_arquivo": pl.nome_arquivo,
                "total_processos": pl.total_processos,
                "origem": pl.origem,
                "subido_legalone": pl.subido_legalone,
                "subido_em": pl.subido_em.isoformat() if pl.subido_em else None,
                "created_at": pl.created_at.isoformat() if pl.created_at else None,
            },
            "progresso": {
                "total": len(rows),
                "cadastrados": cadastrados,
                "pendentes": pendentes,
            },
            "processos": [self._proc_dto(r, nomes) for r in rows],
        }

    def _proc_dto(self, p: BbProcesso, nomes: dict[int, str]) -> dict[str, Any]:
        return {
            "id": p.id,
            "cliente": p.cliente,
            "cnj": p.cnj,
            "npj": p.npj,
            "polo": p.polo,
            "posicao": p.posicao,
            "natureza": p.natureza,
            "acao": p.acao,
            "valor_causa": float(p.valor_causa) if p.valor_causa is not None else None,
            "data_ajuizamento": p.data_ajuizamento,
            "situacao": p.situacao,
            "adverso_principal": p.adverso_principal,
            "responsavel_user_id": p.responsavel_user_id,
            "responsavel_nome": nomes.get(p.responsavel_user_id),
            "escritorio_id": p.escritorio_id,
            "escritorio_path": p.escritorio_path,
            "observacao": p.observacao,
            "status": p.status,
            "planilha_status": p.planilha_status,
            "planilha_id": p.planilha_id,
            "planilha_gerada_em": p.planilha_gerada_em.isoformat() if p.planilha_gerada_em else None,
            "cadastro_confirmado_em": p.cadastro_confirmado_em.isoformat() if p.cadastro_confirmado_em else None,
            "l1_verificado_em": p.l1_verificado_em.isoformat() if p.l1_verificado_em else None,
            "l1_folder": p.l1_folder,
            "ciencia_dada_em": p.ciencia_dada_em.isoformat() if p.ciencia_dada_em else None,
            "l1_lawsuit_id": p.l1_lawsuit_id,
            "erro": p.erro,
            "created_at": p.created_at.isoformat() if p.created_at else None,
        }

    # ── Auditoria de um processo ─────────────────────────────────────────
    def auditoria_processo(self, processo_id: int) -> Optional[dict[str, Any]]:
        p = self.db.get(BbProcesso, processo_id)
        if p is None:
            return None
        nomes = self._mapa_nomes({p.responsavel_user_id})
        envolvidos = [
            {
                "id": e.id,
                "nome": e.nome,
                "papel": e.papel,
                "cpf_cnpj": e.cpf_cnpj,
                "tipo_pessoa": e.tipo_pessoa,
                "status_contato": e.status_contato,
                "l1_contact_id": e.l1_contact_id,
            }
            for e in p.envolvidos
        ]
        eventos = [self._evento_dto(ev) for ev in p.eventos]
        # Envolvidos-de-equipe (derivados da config: equipe do responsável +
        # grupo de ajuizamento atribuído) — como no data.json/aba Envolvidos.
        from app.services.distribuidos_bb.envolvidos_equipe import montar_envolvidos_equipe

        equipe = montar_envolvidos_equipe(self.db, p)
        return {
            "processo": self._proc_dto(p, nomes),
            "envolvidos": envolvidos,
            "envolvidos_equipe": equipe,
            "eventos": eventos,
        }

    # ── Log global de eventos ────────────────────────────────────────────
    def listar_eventos(
        self,
        *,
        secao: Optional[str] = None,
        nivel: Optional[str] = None,
        processo_id: Optional[int] = None,
        run_id: Optional[int] = None,
        busca: Optional[str] = None,
        limit: int = 100,
        offset: int = 0,
    ) -> dict[str, Any]:
        q = self.db.query(BbEvento)
        if secao:
            q = q.filter(BbEvento.secao == secao)
        if nivel:
            q = q.filter(BbEvento.nivel == nivel)
        if processo_id:
            q = q.filter(BbEvento.processo_id == processo_id)
        if run_id:
            q = q.filter(BbEvento.run_id == run_id)
        # Busca por CNJ/NPJ/pasta → auditoria de todo o histórico daquele processo.
        if busca and busca.strip():
            termo = f"%{busca.strip()}%"
            ids = [
                r[0]
                for r in self.db.query(BbProcesso.id)
                .filter(
                    (BbProcesso.cnj.ilike(termo))
                    | (BbProcesso.npj.ilike(termo))
                    | (BbProcesso.adverso_principal.ilike(termo))
                )
                .all()
            ]
            q = q.filter(BbEvento.processo_id.in_(ids or [0]))
        total = q.count()
        rows = q.order_by(BbEvento.id.desc()).limit(limit).offset(offset).all()
        return {"total": total, "items": [self._evento_dto(r) for r in rows]}

    def _evento_dto(self, ev: BbEvento) -> dict[str, Any]:
        return {
            "id": ev.id,
            "run_id": ev.run_id,
            "processo_id": ev.processo_id,
            "secao": ev.secao,
            "acao": ev.acao,
            "nivel": ev.nivel,
            "mensagem": ev.mensagem,
            "dados": ev.dados,
            "created_at": ev.created_at.isoformat() if ev.created_at else None,
        }

    # ── Ingestão do output do RPA legado ─────────────────────────────────
    def ingerir_linhas(
        self, linhas: list[dict[str, Any]], *, run_id: Optional[int] = None
    ) -> dict[str, Any]:
        """Cria/atualiza processos a partir de linhas capturadas e distribui.

        Cada linha aceita as chaves do RPA legado (Processo/CNJ, NPJ, Polo,
        Natureza, Ação, Valor da Causa, Data ajuizamento, Situação, Tramitação,
        Advogado, Adverso Principal). Dedup por fingerprint (CNJ ou NPJ).
        """
        criados = 0
        atualizados = 0
        for i, linha in enumerate(linhas):
            cnj = norm.normalizar_cnj(linha.get("Processo") or linha.get("cnj"))
            npj = (linha.get("NPJ") or linha.get("npj") or "").strip() or None
            fp = norm.fingerprint(cnj, npj)

            existente = (
                self.db.query(BbProcesso).filter(BbProcesso.fingerprint == fp).first()
                if fp != "sem-identidade"
                else None
            )
            proc = existente or BbProcesso(fingerprint=fp, status=PROC_COLETADO)
            proc.run_id = run_id or proc.run_id
            proc.cnj = cnj
            proc.npj = npj
            proc.notificacao_seq = linha.get("Notificação") or linha.get("notificacao_seq")
            proc.polo = (linha.get("Polo") or linha.get("polo") or "").strip() or None
            proc.posicao = norm.polo_para_posicao(proc.polo)
            proc.natureza = (linha.get("Natureza") or linha.get("natureza") or "").strip() or None
            proc.acao = (linha.get("Ação") or linha.get("acao") or "").strip() or None
            proc.valor_causa = norm.parse_valor_causa(linha.get("Valor da Causa") or linha.get("valor_causa"))
            proc.data_ajuizamento = norm.limpar_data_ajuizamento(
                linha.get("Data ajuizamento") or linha.get("data_ajuizamento")
            )
            proc.situacao = (linha.get("Situação") or linha.get("situacao") or "").strip() or None
            proc.tramitacao = (linha.get("Tramitação") or linha.get("tramitacao") or "").strip() or None
            proc.advogado = (linha.get("Advogado") or linha.get("advogado") or "").strip() or None
            proc.adverso_principal = (
                linha.get("Adverso Principal") or linha.get("adverso_principal") or ""
            ).strip() or None
            proc.raw = linha

            if existente is None:
                self.db.add(proc)
                criados += 1
            else:
                atualizados += 1
            self.db.flush()

            registrar_evento(
                self.db,
                secao=SECAO_EXTRACAO,
                nivel=NIVEL_SUCESSO,
                acao="Capturado" if existente is None else "Reatualizado",
                mensagem=(
                    f"Notificação {i + 1} lida: {proc.posicao or '—'} · "
                    f"{proc.natureza or '—'} · {proc.adverso_principal or 'sem adverso'}."
                ),
                dados={
                    "cnj": cnj, "npj": npj, "polo": proc.polo, "natureza": proc.natureza,
                    "valor_causa": float(proc.valor_causa) if proc.valor_causa is not None else None,
                },
                processo_id=proc.id,
                run_id=run_id,
            )

            distribuir_processo(self.db, proc, run_id=run_id)

        self.db.commit()
        return {"criados": criados, "atualizados": atualizados, "total": criados + atualizados}


# ── Exportação xlsx da lista de processos (filtro atual, formatado) ────────
def _montar_xlsx_processos(rows, nomes):
    """Planilha bonitinha com todas as infos dos processos (não é o formato do
    import do L1 — é um export legível pra auditoria/relatório)."""
    import io

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    from app.models.distribuidos_bb import (
        POOL_CADASTRADO_L1,
        POOL_NOVO,
        POOL_PENDENTE_CADASTRO,
    )

    pool_label = {
        POOL_NOVO: "Novo",
        POOL_PENDENTE_CADASTRO: "Pendente cadastro",
        POOL_CADASTRADO_L1: "Cadastrado no L1",
    }
    cols = [
        ("Situação cadastro", 20), ("CNJ", 24), ("NPJ", 18), ("Posição", 11),
        ("Polo", 10), ("Natureza", 14), ("Ação", 28), ("Adverso principal", 30),
        ("Responsável", 26), ("Escritório responsável", 42), ("Observação", 15),
        ("Valor da causa", 16), ("Data ajuizamento", 15), ("Situação (BB)", 26),
        ("Data cadastro L1", 18), ("Pasta L1", 14), ("Capturado em", 18),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Processos"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, (name, w) in enumerate(cols, start=1):
        c = ws.cell(row=1, column=ci, value=name)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 22

    def _dt(x):
        return x.strftime("%d/%m/%Y %H:%M") if x else ""

    for ri, p in enumerate(rows, start=2):
        vals = [
            pool_label.get(p.planilha_status, p.planilha_status),
            p.cnj or "", p.npj or "", p.posicao or "", p.polo or "",
            p.natureza or "", p.acao or "", p.adverso_principal or "",
            nomes.get(p.responsavel_user_id) or "", p.escritorio_path or "",
            p.observacao or "",
            float(p.valor_causa) if p.valor_causa is not None else None,
            p.data_ajuizamento or "", p.situacao or "",
            _dt(p.cadastro_confirmado_em), p.l1_folder or "", _dt(p.created_at),
        ]
        for ci, v in enumerate(vals, start=1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=ci in (7, 8, 9, 10))
            if ci == 12 and v is not None:
                cell.number_format = "R$ #,##0.00"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{max(1, len(rows) + 1)}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

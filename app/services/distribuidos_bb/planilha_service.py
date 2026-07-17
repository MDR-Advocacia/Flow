"""Geração da planilha de migração do Legal One (abas Processos + Envolvidos).

Porta o `gerar_planilha.py` legado, mas puxando de `bbd_processos` (coletados +
distribuídos) e da config EDITÁVEL (escritórios, valores padrão, equipes,
grupos de ajuizamento). O operador exporta e importa no L1 — o import dispara
o workflow nativo (que o POST /Lawsuits da API REST não dispara).

Colunas (0-indexadas) da aba Processos seguem o MODELO LEGAL ONE.xlsx.
"""
from __future__ import annotations

import io
from copy import copy
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional
from zoneinfo import ZoneInfo

import openpyxl
from sqlalchemy.orm import Session

from app.models.distribuidos_bb import (
    NIVEL_AVISO,
    PLANILHA_MANUAL,
    POOL_NOVO,
    POOL_PENDENTE_CADASTRO,
    PROC_DISTRIBUIDO,
    SECAO_PLANILHA,
    BbConfig,
    BbEnvolvido,
    BbPlanilha,
    BbProcesso,
)
from app.models.legal_one import LegalOneUser
from app.services.distribuidos_bb import normalizacao as norm
from app.services.distribuidos_bb.cadastro_l1 import parse_tramitacao
from app.services.distribuidos_bb.envolvidos_equipe import montar_envolvidos_equipe
from app.services.distribuidos_bb.log_service import registrar_evento

_TEMPLATE = Path(__file__).parent / "templates" / "MODELO_LEGAL_ONE.xlsx"
_TZ_BR = ZoneInfo("America/Sao_Paulo")


def _cfg(db: Session, chave: str, default: str = "") -> str:
    c = db.get(BbConfig, chave)
    return c.valor if (c and c.valor is not None) else default


def gerar_planilha(
    db: Session,
    *,
    processo_ids: Optional[list[int]] = None,
    status: Optional[str] = "DISTRIBUIDO",
) -> tuple[io.BytesIO, int]:
    """Monta o xlsx a partir dos processos. Devolve (BytesIO, total_processos)."""
    q = db.query(BbProcesso)
    if processo_ids:
        q = q.filter(BbProcesso.id.in_(processo_ids))
    elif status:
        q = q.filter(BbProcesso.status == status)
    processos = q.order_by(BbProcesso.id).all()

    wb = openpyxl.load_workbook(_TEMPLATE)
    ws_p = wb["Processos"]
    ws_e = wb["Envolvidos"]

    # Estilo da linha-modelo (row 2) e limpeza (remove exemplo + dados antigos)
    modelo_p = list(ws_p.iter_rows(min_row=2, max_row=2))[0]
    modelo_e = list(ws_e.iter_rows(min_row=2, max_row=2))[0]
    estilo_p = [c._style if c.has_style else None for c in modelo_p]
    estilo_e = [c._style if c.has_style else None for c in modelo_e]
    if ws_p.max_row >= 2:
        ws_p.delete_rows(2, ws_p.max_row)
    if ws_e.max_row >= 2:
        ws_e.delete_rows(2, ws_e.max_row)

    # Constantes editáveis (Valores Padrão)
    tipo_registro = _cfg(db, "tipo_registro", "Processo")
    tipo = _cfg(db, "tipo", "Judicial")
    status_v = _cfg(db, "status", "Ativo")
    origem = _cfg(db, "escritorio_origem", "MDR Advocacia")
    situacao_env = _cfg(db, "situacao_envolvido", "Outros")

    # Cliente por linha (BB × Ativos): nome/CNPJ/tipo do dono do processo.
    def _cliente_cfg(cli: str) -> tuple[str, str, str]:
        if cli == "ATIVOS":
            return (
                _cfg(db, "ativos_cliente_nome", "Ativos S.A. Securitizadora de Créditos Financeiros"),
                _cfg(db, "ativos_cliente_cpf_cnpj", "05.437.257/0001-29"),
                _cfg(db, "ativos_cliente_tipo", "PJ"),
            )
        return (
            _cfg(db, "cliente_nome", "Banco do Brasil S.A."),
            _cfg(db, "cliente_cpf_cnpj", "00.000.000/0001-91"),
            _cfg(db, "cliente_tipo", "PJ"),
        )

    # Nomes dos responsáveis
    resp_ids = {p.responsavel_user_id for p in processos if p.responsavel_user_id}
    nomes = dict(
        db.query(LegalOneUser.id, LegalOneUser.name)
        .filter(LegalOneUser.id.in_(resp_ids or {0}))
        .all()
    )

    def _aplica_estilo(ws, row, estilos):
        for ci, st in enumerate(estilos, start=1):
            if st is not None:
                ws.cell(row=row, column=ci)._style = copy(st)

    row_p = 2
    row_e = 2
    chave = 1
    for p in processos:
        tram = parse_tramitacao(p.tramitacao)
        cliente_nome, cliente_cpf, cliente_tipo = _cliente_cfg(p.cliente)
        if p.cliente_nome:  # pasta avulsa: cliente informado no modal vence a config
            cliente_nome = p.cliente_nome
            cliente_cpf = p.cliente_cpf_cnpj or ""
            cliente_tipo = p.cliente_tipo or "PJ"
        doc_cliente = norm.apenas_digitos(cliente_cpf)
        # Contrário principal (1º envolvido com doc que NÃO seja o cliente)
        adverso_cpf = adverso_tipo = None
        for e in db.query(BbEnvolvido).filter(BbEnvolvido.processo_id == p.id).all():
            if e.cpf_cnpj and norm.apenas_digitos(e.cpf_cnpj) != doc_cliente:
                adverso_cpf, adverso_tipo = e.cpf_cnpj, e.tipo_pessoa
                break

        linha = [None] * 31
        linha[0] = chave
        linha[2] = tipo_registro
        linha[3] = p.cnj
        linha[4] = tipo
        linha[5] = status_v
        linha[6] = p.natureza
        linha[7] = cliente_nome
        linha[8] = p.posicao
        linha[9] = cliente_cpf
        linha[10] = cliente_tipo
        linha[11] = p.adverso_principal
        linha[12] = adverso_cpf
        linha[13] = adverso_tipo
        linha[15] = p.data_ajuizamento
        linha[16] = p.acao
        linha[17] = p.npj if p.cliente == "BB" else ""
        linha[19] = tram["uf"]
        linha[20] = tram["cidade"]
        linha[21] = tram["orgao"]
        linha[24] = float(p.valor_causa) if p.valor_causa is not None else None
        linha[26] = nomes.get(p.responsavel_user_id)
        linha[27] = p.escritorio_path
        linha[28] = origem
        linha[29] = p.observacao

        _aplica_estilo(ws_p, row_p, estilo_p)
        for ci in range(1, 32):
            ws_p.cell(row=row_p, column=ci, value=linha[ci - 1])
        row_p += 1

        # Envolvidos (equipe do responsável + grupo de ajuizamento)
        for env in montar_envolvidos_equipe(db, p):
            _aplica_estilo(ws_e, row_e, estilo_e)
            ws_e.cell(row=row_e, column=1, value=chave)
            ws_e.cell(row=row_e, column=2, value=env["nome"])
            ws_e.cell(row=row_e, column=5, value=situacao_env)
            ws_e.cell(row=row_e, column=6, value=env["classificacao"])
            row_e += 1

        chave += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, len(processos)


def contar_pool_novos(db: Session, *, cliente: Optional[str] = None) -> int:
    """Quantos processos estão no pool aguardando planilha (NOVO + distribuídos).

    `cliente` restringe a contagem (o pool é por cliente: o operador do BB não
    deve ver os Ativos pendentes somados no total dele).
    """
    q = db.query(BbProcesso).filter(
        BbProcesso.planilha_status == POOL_NOVO,
        BbProcesso.status == PROC_DISTRIBUIDO,
    )
    if cliente:
        q = q.filter(BbProcesso.cliente == cliente)
    return q.count()


def gerar_e_persistir(
    db: Session,
    *,
    processo_ids: Optional[list[int]] = None,
    origem: str = PLANILHA_MANUAL,
    cliente: Optional[str] = None,
) -> Optional[BbPlanilha]:
    """Gera a planilha do POOL e a arquiva em `bbd_planilhas`. Devolve a linha.

    Regra do pool: pega TODOS os processos marcados NOVO (e já distribuídos),
    gera a planilha e os marca como PLANILHA_GERADA (vinculando à planilha).
    A próxima coleta traz os processos novos como NOVO de novo.

    - Se `processo_ids` for dado, usa exatamente esses (subset manual).
    - Devolve `None` quando não há nada no pool (não cria planilha vazia).
    - NÃO faz commit — quem chama controla a transação.
    """
    if processo_ids is None:
        q = db.query(BbProcesso).filter(
            BbProcesso.planilha_status == POOL_NOVO,
            BbProcesso.status == PROC_DISTRIBUIDO,
        )
        # Sem isto o pool é comum: o auto-cadastro da coleta do BB varreria junto
        # os processos do Ativos que estivessem pendentes (misturando clientes na
        # mesma planilha e no relatório do run).
        if cliente:
            q = q.filter(BbProcesso.cliente == cliente)
        processos = q.order_by(BbProcesso.id).all()
    else:
        processos = (
            db.query(BbProcesso).filter(BbProcesso.id.in_(processo_ids or [0])).all()
        )
    if not processos:
        return None

    # TRAVA: não cadastrar processo sem responsável. Este é o ponto por onde os
    # três caminhos passam (coleta do BB, lote do Ativos e o botão manual), então
    # a regra vale pra todos. Antes de excluir, TENTA distribuir de novo — a fila
    # pode ter sido configurada depois da ingestão (caso clássico do Ativos, cujos
    # escritórios nascem com a fila vazia). Quem continuar sem responsável fica no
    # pool como NOVO e não entra na planilha.
    from app.services.distribuidos_bb.distribuicao_service import distribuir_processo

    elegiveis: list[BbProcesso] = []
    sem_responsavel: list[BbProcesso] = []
    for p in processos:
        if p.responsavel_user_id is None:
            distribuir_processo(db, p)
        (elegiveis if p.responsavel_user_id is not None else sem_responsavel).append(p)

    if sem_responsavel:
        filas = sorted({(p.escritorio_path or p.posicao or "—") for p in sem_responsavel})
        registrar_evento(
            db,
            secao=SECAO_PLANILHA,
            nivel=NIVEL_AVISO,
            acao="Sem responsável — fora da planilha",
            mensagem=(
                f"{len(sem_responsavel)} processo(s) ficaram FORA da planilha por não terem "
                f"responsável: a fila destes escritórios está vazia — {', '.join(filas)}. "
                f"Eles seguem no pool; configure a fila e gere a planilha de novo."
            ),
            dados={"quantidade": len(sem_responsavel), "escritorios": filas},
        )

    if not elegiveis:
        return None
    processos = elegiveis

    ids = [p.id for p in processos]
    buf, total = gerar_planilha(db, processo_ids=ids, status=None)
    if total == 0:
        return None

    dados = buf.getvalue()
    carimbo = datetime.now(_TZ_BR).strftime("%Y%m%d_%H%M")
    # O nome carrega o cliente (a planilha do Ativos não é "DISTRIBUIDOS_BB").
    # Quando os ids vêm soltos, deduz pelo conteúdo; homogêneo é o caso normal.
    cli = cliente or (processos[0].cliente if processos else None)
    if cli and any(p.cliente != cli for p in processos):
        rotulo = "MISTA"
    else:
        rotulo = {"ATIVOS": "ATIVOS", "OUTRO": "AVULSA"}.get(cli or "", "DISTRIBUIDOS_BB")
    nome = f"PLANILHA_MIGRACAO_{rotulo}_{carimbo}.xlsx"

    planilha = BbPlanilha(
        nome_arquivo=nome,
        conteudo=dados,
        total_processos=total,
        tamanho_bytes=len(dados),
        origem=origem,
        status_origem=PROC_DISTRIBUIDO,
    )
    db.add(planilha)
    db.flush()   # garante planilha.id

    # Marca o pool como PENDENTE_CADASTRO (aguardando o L1) e vincula à planilha.
    agora = datetime.now(timezone.utc)
    for p in processos:
        p.planilha_status = POOL_PENDENTE_CADASTRO
        p.planilha_id = planilha.id
        p.planilha_gerada_em = agora

    return planilha

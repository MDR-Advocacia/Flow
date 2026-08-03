"""Intake do Sistema de Encerramentos: encerrar processo no Legal One.

O Sistema de Encerramentos (MDR) chama este endpoint quando o operador
encerra um processo lá — o Flow executa o encerramento no Legal One via
API oficial. Autenticado por API key no header `X-Encerramentos-Api-Key`
(env `ENCERRAMENTOS_INTAKE_API_KEY`), SEM JWT — mesmo padrão dos intakes
de Prazos Iniciais, Classificador e OneRequest. Registrado sem
`protected_dependencies` em main.py.

TODA chamada (sucesso ou falha) é gravada em `encerramentos_l1_intake`
para o menu "Encerramentos" da UI (router protegido `router` abaixo):
a gestão enxerga o que está sendo encerrado via integração, por quem e
com qual desfecho.

Regras de negócio (decididas com a operação em 30/07/2026):
- `closingReason` (MotivoEncerramento no L1) recebe o valor pronto vindo
  do Encerramentos: NOME COMPLETO de quem encerrou + data e hora
  (convenção que a equipe já usava manualmente no fluxo da UI).
- O responsável da pasta NÃO é alterado.
- Campos de resultado (result/resultType/resultReason/datas) não são
  tocados — seguem preenchidos pela equipe no L1.

Contrato completo: docs/integracao-flow-legalone.md no repo Encerramentos.
"""

import logging
from typing import Optional

import requests
from fastapi import APIRouter, Depends, Header, HTTPException, Query, status
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session

from app.core.auth import require_permission
from app.core.config import settings
from app.core.dependencies import get_api_client, get_db
from app.models.legal_one import EncerramentoL1Intake, LegalOneUser
from app.services.legal_one_client import LegalOneApiClient

logger = logging.getLogger(__name__)

intake_router = APIRouter(prefix="/legalone", tags=["Encerramentos (Intake)"])
router = APIRouter(prefix="/legalone", tags=["Encerramentos"])

# Rotulos exibidos na UI e no Excel (mesma fonte, sem divergir)
STATUS_LABELS = {
    "ok": "Encerrado",
    "ja_encerrado": "Já estava encerrado",
    "nao_encontrado": "CNJ não encontrado",
    "conflito": "Conflito",
    "erro_l1": "Erro no Legal One",
}


def _filtrar(query, status_filtro: Optional[str], q: Optional[str]):
    """Filtros compartilhados por listagem e exportação — o Excel tem que
    sair exatamente com o recorte que a tela está mostrando."""
    if status_filtro:
        query = query.filter(EncerramentoL1Intake.status == status_filtro)
    if q:
        like = f"%{q.strip()}%"
        query = query.filter(
            EncerramentoL1Intake.numero_cnj.ilike(like)
            | EncerramentoL1Intake.operador_nome.ilike(like)
        )
    return query


def _validate_intake_api_key(
    x_encerramentos_api_key: Optional[str] = Header(
        default=None, alias="X-Encerramentos-Api-Key"
    ),
) -> str:
    """
    Autentica o Sistema de Encerramentos por header `X-Encerramentos-Api-Key`.

    Aceita múltiplas chaves em `ENCERRAMENTOS_INTAKE_API_KEY` (separadas por
    vírgula) pra rotação sem downtime. Se nenhuma chave estiver configurada,
    o endpoint fica explicitamente bloqueado (503) — evita rota aberta em
    produção por esquecimento de config.
    """
    valid_keys = settings.encerramentos_intake_api_keys
    if not valid_keys:
        logger.error(
            "ENCERRAMENTOS_INTAKE_API_KEY não configurada — encerramento rejeitado."
        )
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="Endpoint de encerramento não configurado.",
        )
    if not x_encerramentos_api_key or x_encerramentos_api_key not in valid_keys:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="API key inválida ou ausente no header X-Encerramentos-Api-Key.",
        )
    return x_encerramentos_api_key


class EncerramentoPayload(BaseModel):
    numero_cnj: str = Field(..., min_length=10, description="CNJ do processo")
    data_encerramento: str = Field(
        ..., description="Data do encerramento (yyyy-mm-dd)"
    )
    motivo_encerramento: str = Field(
        ...,
        min_length=1,
        description="Valor pronto pro MotivoEncerramento do L1 "
        "(NOME COMPLETO - dd/mm/aaaa hh:mm)",
    )
    # Contexto/auditoria (não escritos no L1 diretamente)
    origem: str = Field(default="encerramentos")
    encerrado_em: Optional[str] = None
    operador_nome: Optional[str] = None
    operador_email: Optional[str] = None
    justificativa: Optional[str] = None


def _campos_faltando(body: str) -> str:
    """Extrai os nomes de campo do erro de validação do L1, em formato legível.

    O L1 devolve o nome interno do custom field
    (ex.: `numeroDoCliente_ProcessoEntitySchema_p3687_o`); aqui fica só a
    parte útil pra operação saber o que preencher no cadastro.
    """
    import json as _json
    import re as _re

    try:
        detalhes = _json.loads(body).get("error", {}).get("details", [])
    except Exception:
        return ""
    nomes = []
    for d in detalhes:
        alvo = str(d.get("target") or "").split("_")[0]
        if alvo:
            # camelCase -> "Camel Case"
            nomes.append(_re.sub(r"(?<!^)(?=[A-Z])", " ", alvo).strip().capitalize())
    return ", ".join(dict.fromkeys(nomes))


def _registrar(
    db: Session,
    payload: EncerramentoPayload,
    status_registro: str,
    lawsuit_id: Optional[int] = None,
    detalhe: str = "",
) -> None:
    """Grava o rastro da chamada (nunca derruba o fluxo se falhar)."""
    try:
        db.add(EncerramentoL1Intake(
            numero_cnj=payload.numero_cnj,
            lawsuit_id=lawsuit_id,
            status=status_registro,
            data_encerramento=payload.data_encerramento,
            motivo_encerramento=payload.motivo_encerramento,
            operador_nome=payload.operador_nome,
            operador_email=payload.operador_email,
            justificativa=payload.justificativa,
            origem=payload.origem,
            detalhe=detalhe[:2000] if detalhe else None,
        ))
        db.commit()
    except Exception:  # pragma: no cover - rastro nunca pode matar o encerramento
        logger.exception("Falha gravando rastro do encerramento %s", payload.numero_cnj)
        db.rollback()


@intake_router.post("/encerramento")
def encerrar_processo_legalone(
    payload: EncerramentoPayload,
    api_key: str = Depends(_validate_intake_api_key),
    client: LegalOneApiClient = Depends(get_api_client),
    db: Session = Depends(get_db),
):
    """
    Encerra o processo no Legal One (closed + closingDate + closingReason).

    Respostas:
    - 200 {status: "ok"|"ja_encerrado", lawsuit_id}
    - 404 CNJ sem correspondência no L1
    - 409 já encerrado no L1 com dados divergentes dos enviados
    - 502 falha na API do Legal One
    """
    lawsuit = client.search_lawsuit_by_cnj(payload.numero_cnj)
    if not lawsuit:
        _registrar(db, payload, "nao_encontrado")
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Processo {payload.numero_cnj} não encontrado no Legal One.",
        )
    lawsuit_id = int(lawsuit["id"])

    # Idempotência: retry do Encerramentos não pode virar erro nem sobrescrever
    try:
        atual = client.get_lawsuit_by_id(
            lawsuit_id, params={"$select": "id,closed,closingDate,closingReason"}
        )
    except requests.exceptions.HTTPError as exc:
        logger.error("Falha lendo lawsuit %s antes do encerramento: %s", lawsuit_id, exc)
        _registrar(db, payload, "erro_l1", lawsuit_id, f"Falha lendo lawsuit: {exc}")
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail="Falha ao consultar o processo no Legal One.",
        ) from exc

    if atual.get("closed"):
        reason_atual = (atual.get("closingReason") or "").strip()
        if reason_atual == payload.motivo_encerramento.strip():
            logger.info(
                "Lawsuit %s já encerrado com o mesmo motivo — idempotente.",
                lawsuit_id,
            )
            _registrar(db, payload, "ja_encerrado", lawsuit_id)
            return {"status": "ja_encerrado", "lawsuit_id": lawsuit_id}
        detalhe = (
            "Processo já encerrado no Legal One com dados divergentes "
            f"(closingReason atual: '{reason_atual}' | "
            f"closingDate: {atual.get('closingDate')})."
        )
        _registrar(db, payload, "conflito", lawsuit_id, detalhe)
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail=detalhe)

    try:
        resultado = client.close_lawsuit(
            lawsuit_id=lawsuit_id,
            closing_date=payload.data_encerramento,
            closing_reason=payload.motivo_encerramento,
        )
    except requests.exceptions.HTTPError as exc:
        body = exc.response.text[:600] if exc.response is not None else str(exc)
        codigo = exc.response.status_code if exc.response is not None else 0
        logger.error("L1 recusou o encerramento do lawsuit %s: %s", lawsuit_id, body)

        # Validação do L1 (400) = pendência de CADASTRO no processo, não falha
        # transitória: repetir dá o mesmo erro. Devolve 409 pra origem marcar
        # DIVERGENTE (ação humana) em vez de ERRO com retry em loop.
        # Caso típico: campos customizados obrigatórios vazios no L1
        # (ex.: "numeroDoCliente" / NPJ e "dataDeTerceirizacaoRecebimento").
        if codigo == 400:
            faltando = _campos_faltando(body)
            detalhe = (
                "Legal One recusou por validação de cadastro"
                + (f" — campo(s) obrigatório(s) sem preenchimento no L1: {faltando}." if faltando else f": {body}")
            )
            _registrar(db, payload, "conflito", lawsuit_id, detalhe)
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail=detalhe) from exc

        _registrar(db, payload, "erro_l1", lawsuit_id, body)
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail=f"Legal One recusou o encerramento: {body}",
        ) from exc

    logger.info(
        "Lawsuit %s (%s) encerrado via Encerramentos por %s.",
        lawsuit_id, payload.numero_cnj, payload.operador_nome or "?",
    )
    _registrar(db, payload, "ok", lawsuit_id)
    return {"status": "ok", "lawsuit_id": lawsuit_id, "entity": resultado.get("entity")}


# ── Listagem para a UI (menu "Encerramentos") ─────────────────────────


@router.get("/encerramentos")
def listar_encerramentos(
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=25, ge=1, le=100),
    status_filtro: Optional[str] = Query(default=None, alias="status"),
    q: Optional[str] = Query(default=None, description="Busca por CNJ ou operador"),
    db: Session = Depends(get_db),
    current_user: LegalOneUser = Depends(require_permission("encerramentos")),
):
    """Listagem paginada (regra da casa) do rastro de encerramentos.

    Acesso pela permissão de módulo `can_use_encerramentos` (admins bypassam),
    no mesmo padrão de Publicações / Prazos Iniciais / OneRequest."""
    query = _filtrar(db.query(EncerramentoL1Intake), status_filtro, q)

    total = query.count()
    itens = (
        query.order_by(EncerramentoL1Intake.created_at.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    # Contadores por status (do recorte de busca, ignorando o filtro de status)
    base = _filtrar(db.query(EncerramentoL1Intake), None, q)
    contadores = {s: 0 for s in ("ok", "ja_encerrado", "nao_encontrado", "conflito", "erro_l1")}
    from sqlalchemy import func as safunc

    for st, n in base.with_entities(
        EncerramentoL1Intake.status, safunc.count(EncerramentoL1Intake.id)
    ).group_by(EncerramentoL1Intake.status):
        contadores[st] = n

    return {
        "items": [
            {
                "id": i.id,
                "created_at": i.created_at.isoformat() if i.created_at else None,
                "numero_cnj": i.numero_cnj,
                "lawsuit_id": i.lawsuit_id,
                "status": i.status,
                "data_encerramento": i.data_encerramento,
                "motivo_encerramento": i.motivo_encerramento,
                "operador_nome": i.operador_nome,
                "operador_email": i.operador_email,
                "justificativa": i.justificativa,
                "origem": i.origem,
                "detalhe": i.detalhe,
            }
            for i in itens
        ],
        "total": total,
        "page": page,
        "page_size": page_size,
        "contadores": contadores,
    }


@router.get("/encerramentos/export", summary="Exporta os encerramentos filtrados em Excel (xlsx)")
def exportar_encerramentos(
    status_filtro: Optional[str] = Query(default=None, alias="status"),
    q: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
    current_user: LegalOneUser = Depends(require_permission("encerramentos")),
):
    """Excel com os MESMOS filtros da listagem — insumo pra corrigir o
    cadastro no L1 e reprocessar. Sem paginação: sai tudo do recorte."""
    from datetime import datetime as _dt
    from io import BytesIO

    from fastapi import Response
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    itens = _filtrar(db.query(EncerramentoL1Intake), status_filtro, q).order_by(
        EncerramentoL1Intake.created_at.desc()
    ).all()

    colunas = [
        ("Recebido em", 18, lambda i: i.created_at.strftime("%d/%m/%Y %H:%M") if i.created_at else ""),
        ("CNJ", 26, lambda i: i.numero_cnj),
        ("Lawsuit (L1)", 12, lambda i: i.lawsuit_id),
        ("Desfecho", 22, lambda i: STATUS_LABELS.get(i.status, i.status)),
        ("Dt. Encerramento", 16, lambda i: i.data_encerramento or ""),
        ("Motivo (L1)", 42, lambda i: i.motivo_encerramento or ""),
        ("Operador", 30, lambda i: i.operador_nome or ""),
        ("E-mail", 30, lambda i: i.operador_email or ""),
        ("Justificativa", 45, lambda i: i.justificativa or ""),
        ("Origem", 14, lambda i: i.origem or ""),
        ("Detalhe / motivo da recusa", 90, lambda i: i.detalhe or ""),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Encerramentos L1"
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="1F4E79")
    for col, (titulo, largura, _g) in enumerate(colunas, start=1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.font = head_font
        cell.fill = head_fill
        ws.column_dimensions[get_column_letter(col)].width = largura
    for row, item in enumerate(itens, start=2):
        for col, (_t, _w, getter) in enumerate(colunas, start=1):
            ws.cell(row=row, column=col, value=getter(item))
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    nome = f"encerramentos-legalone-{_dt.now().strftime('%Y%m%d-%H%M')}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nome}"'},
    )

"""Endpoints do Balanceador de Agenda.

Leitura: diagnóstico de carga + matriz de redistribuição + detalhe por subtipo
(snapshot perf_l1_tarefa) + fila AO VIVO do L1. Reusa o gate por time do Minha
Equipe.

Escrita REAL (2026-07-02): `/reatribuir` dispara um job server-backed que troca
responsável+executante no L1 (PATCH normal; tarefa de Workflow vai pro bucket
`workflow_bloqueadas`). Progresso via `/reatribuir/status`. Ver
app/services/performance/reatribuir_job.py.
"""

from fastapi import APIRouter, Depends, HTTPException, Query
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.api.v1.endpoints.performance import require_team_access
from app.core.auth import get_current_user
from app.core.dependencies import get_db
from app.models.legal_one import LegalOneUser
from app.services.performance.balanceador import BalanceadorService

router = APIRouter(prefix="/balanceador", tags=["Balanceador de Agenda"])
_team = Depends(require_team_access)


@router.get("/diagnostico", summary="Carga pendente por colaborador (atrasado/fatal hoje/futuro)", dependencies=[_team])
def diagnostico(team: str = Query(...), db: Session = Depends(get_db)):
    return {"colaboradores": BalanceadorService(db).diagnostico(team)}


@router.get("/redistribuir", summary="Matriz subtipo × colaborador dos escolhidos", dependencies=[_team])
def redistribuir(
    team: str = Query(...),
    pessoas: str = Query(..., description="ids separados por vírgula"),
    dias: int = Query(0, description="janela futura em dias; 0 = tudo"),
    db: Session = Depends(get_db),
):
    ids = [int(x) for x in pessoas.split(",") if x.strip().isdigit()]
    return {"matriz": BalanceadorService(db).redistribuir_matriz(team, ids, dias)}


@router.get("/tarefas", summary="Tarefas individuais de um (colaborador, subtipo)", dependencies=[_team])
def tarefas(
    team: str = Query(...),
    pessoa_id: int = Query(...),
    subtipo: str = Query(...),
    dias: int = Query(0),
    db: Session = Depends(get_db),
):
    return {"tarefas": BalanceadorService(db).redistribuir_tarefas(team, pessoa_id, subtipo, dias)}


@router.get("/descricoes", summary="Descrição (assunto) ao vivo do L1 por task ids", dependencies=[_team])
def descricoes(team: str = Query(...), ids: str = Query(...), db: Session = Depends(get_db)):
    idlist = [int(x) for x in ids.split(",") if x.strip().isdigit()]
    return {"descricoes": BalanceadorService(db).descricoes(idlist)}


@router.get("/live-pessoa", summary="Pendentes AO VIVO do L1 de uma pessoa (matriz + detalhe)", dependencies=[_team])
def live_pessoa(
    team: str = Query(...),
    pessoa_id: int = Query(...),
    dias: int = Query(0, description="legado: janela de N dias (usar inicio/fim)"),
    incluir_atrasadas: bool = Query(True),
    inicio: str | None = Query(None, description="faixa exata: data de conclusão prevista inicial (YYYY-MM-DD)"),
    fim: str | None = Query(None, description="faixa exata: data de conclusão prevista final (YYYY-MM-DD)"),
    db: Session = Depends(get_db),
):
    return BalanceadorService(db).live_pessoa(
        team, pessoa_id, dias, incluir_atrasadas, inicio=inicio, fim=fim,
    )


@router.get("/usuarios", summary="Destinos da fila: setor primeiro + busca externa no L1", dependencies=[_team])
def usuarios(team: str = Query(...), busca: str = Query(""), db: Session = Depends(get_db)):
    return {"usuarios": BalanceadorService(db).buscar_usuarios(team, busca)}


@router.get("/fila-pref", summary="Destinos recorrentes da fila p/ (origem, subtipo)", dependencies=[_team])
def fila_pref_sugestoes(
    team: str = Query(...),
    origem_pessoa_id: int = Query(...),
    subtipo: str = Query(...),
    db: Session = Depends(get_db),
):
    return {"sugestoes": BalanceadorService(db).sugestoes_fila(team, origem_pessoa_id, subtipo)}


class FilaPrefReq(BaseModel):
    origem_pessoa_id: int
    subtipo: str
    alvos: list  # [{id, nome}]


@router.post("/fila-pref", summary="Aprende os destinos da fila (origem, subtipo)", dependencies=[_team])
def fila_pref_registrar(team: str = Query(...), req: FilaPrefReq = ..., db: Session = Depends(get_db)):
    return BalanceadorService(db).registrar_fila_pref(team, req.origem_pessoa_id, req.subtipo, req.alvos)


class RegistrarLogReq(BaseModel):
    movimentos: list


@router.post("/log", summary="Registra o log de uma redistribuição (aba Relatórios)", dependencies=[_team])
def registrar_log(
    team: str = Query(...),
    req: RegistrarLogReq = ...,
    current_user: LegalOneUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    return BalanceadorService(db).registrar_log(team, current_user, req.movimentos)


@router.get("/logs", summary="Lista os logs de redistribuição do time (paginado)", dependencies=[_team])
def listar_logs(
    team: str = Query(...),
    limit: int = Query(10, ge=1, le=100),
    offset: int = Query(0, ge=0),
    db: Session = Depends(get_db),
):
    return BalanceadorService(db).listar_logs(team, limit=limit, offset=offset)


# ── Reatribuição EM LOTE (escrita REAL no L1) — job server-backed com progresso ──


class ReatribuirItem(BaseModel):
    task_id: int
    to_id: int | None = None
    to_nome: str | None = None
    origem: str | None = "tarefa"  # "tarefa" (/Tasks) | "compromisso" (/Appointments)


class ReatribuirReq(BaseModel):
    itens: list[ReatribuirItem]
    movimentos: list = []  # move-level (from/to/subtipo/qtd), só pra auditoria


@router.post(
    "/reatribuir",
    summary="Dispara a reatribuição EM LOTE das tarefas (PATCH normal; Workflow vai pro bucket bloqueado)",
    dependencies=[_team],
)
def reatribuir(
    team: str = Query(...),
    dry_run: bool = Query(False, description="só lê participantes (prova acesso, conta) — não grava"),
    req: ReatribuirReq = ...,
    current_user: LegalOneUser = Depends(get_current_user),
):
    from app.services.performance import reatribuir_job

    itens = [i.model_dump() for i in req.itens]
    job_id = reatribuir_job.iniciar(team, itens, req.movimentos, dry_run, current_user)
    return {"job_id": job_id, "status": reatribuir_job.status(job_id)}


@router.get("/reatribuir/status", summary="Progresso do job de reatribuição", dependencies=[_team])
def reatribuir_status(team: str = Query(...), job_id: str = Query(...)):
    from app.services.performance import reatribuir_job

    st = reatribuir_job.status(job_id)
    if st is None:
        return {"status": "not_found"}
    return st


@router.post("/reatribuir/abort", summary="Aborta o job de reatribuição", dependencies=[_team])
def reatribuir_abort(team: str = Query(...), job_id: str = Query(...)):
    from app.services.performance import reatribuir_job

    return {"ok": reatribuir_job.solicitar_abort(job_id)}


# Motivo legível por tarefa (mapa único usado no Excel; o front tem o espelho).
_REASON_LABELS = {
    "reassigned": "Reatribuída (API)",
    "reassigned_web": "Reatribuída (caminho web)",
    "dry_ok": "Simulada — leitura OK",
    "web_pendente": "Aguardando caminho web",
    "web_erro": "Erro no caminho web",
    "web_nao_refletiu": "Web não confirmou a troca",
    "web_sem_papel_atual": "Sem titular atual do papel (manual)",
    "workflow_locked": "Workflow — travada na API (manual)",
    "workflow_web_pendente": "Aguardando caminho web",
    "workflow_web_erro": "Erro no caminho web",
    "workflow_web_nao_refletiu": "Web não confirmou a troca",
    "workflow_sem_papel_atual": "Sem titular atual do papel (manual)",
    "destino_nao_resolvido": "Destino não resolvido no L1",
    "error": "Erro na API",
}


@router.get(
    "/reatribuir/jobs",
    summary="Execuções de reatribuição do time (em andamento + histórico, paginado)",
    dependencies=[_team],
)
def reatribuir_jobs(
    team: str = Query(...),
    limit: int = Query(10, ge=1, le=100),
    offset: int = Query(0, ge=0),
    db: Session = Depends(get_db),
):
    from app.models.performance import BalanceadorReatribuirJob

    base = db.query(BalanceadorReatribuirJob).filter(BalanceadorReatribuirJob.team == team)
    total = base.count()
    rows = (
        base.order_by(BalanceadorReatribuirJob.iniciado_em.desc())
        .offset(offset)
        .limit(limit)
        .all()
    )
    return {
        "total": total,
        "items": [
            {
                "job_id": j.id,
                "status": j.status,
                "dry_run": j.dry_run,
                "total": j.total or 0,
                "feito": j.feito or 0,
                "reatribuidas": j.reatribuidas or 0,
                "workflow_bloqueadas": j.workflow_bloqueadas or 0,
                "falhas": j.falhas or 0,
                "criado_por_nome": j.criado_por_nome,
                "iniciado_em": j.iniciado_em.isoformat() if j.iniciado_em else None,
                "terminado_em": j.terminado_em.isoformat() if j.terminado_em else None,
            }
            for j in rows
        ],
    }


@router.get(
    "/reatribuir/jobs/{job_id}/detalhe",
    summary="Detalhe por tarefa de uma execução (motivo legível)",
    dependencies=[_team],
)
def reatribuir_job_detalhe(job_id: str, team: str = Query(...), db: Session = Depends(get_db)):
    from sqlalchemy import text as _text

    from app.models.performance import BalanceadorReatribuirJob

    j = db.get(BalanceadorReatribuirJob, job_id)
    if not j or j.team != team:
        return {"tarefas": []}
    detalhe = j.detalhe or []
    # Enriquece com subtipo/pasta/cnj do snapshot (quando a tarefa está nele).
    ids = [int(d["task_id"]) for d in detalhe if d.get("task_id")]
    info: dict = {}
    if ids:
        for r in db.execute(
            _text(
                "SELECT DISTINCT ON (l1_task_id) l1_task_id, subtipo, pasta, cnj "
                "FROM perf_l1_tarefa WHERE l1_task_id = ANY(:ids)"
            ),
            {"ids": ids},
        ).fetchall():
            info[int(r.l1_task_id)] = {"subtipo": r.subtipo, "pasta": r.pasta, "cnj": r.cnj}
    return {
        "tarefas": [
            {
                **d,
                "resultado": _REASON_LABELS.get(d.get("reason"), d.get("reason")),
                **(info.get(int(d["task_id"])) if d.get("task_id") and int(d["task_id"]) in info else {}),
            }
            for d in detalhe
        ]
    }


# Reasons que contam como SUCESSO — o resto (falha/pendência/manual) é
# retentável. dry_ok não conta: simulação não gravou nada pra "refazer".
_REASON_OK = {"reassigned", "reassigned_web", "dry_ok"}


@router.post(
    "/reatribuir/jobs/{job_id}/retry",
    summary="Refaz só as tarefas que falharam/ficaram pendentes numa execução",
    dependencies=[_team],
)
def reatribuir_job_retry(
    job_id: str,
    team: str = Query(...),
    db: Session = Depends(get_db),
    current_user: LegalOneUser = Depends(get_current_user),
):
    from app.models.performance import BalanceadorReatribuirJob
    from app.services.performance import reatribuir_job

    j = db.get(BalanceadorReatribuirJob, job_id)
    if not j or j.team != team:
        raise HTTPException(status_code=404, detail="Execução não encontrada")
    if j.status != "done":
        raise HTTPException(status_code=409, detail="A execução ainda está em andamento")
    if j.dry_run:
        raise HTTPException(status_code=400, detail="Simulação não grava no L1 — nada a refazer")

    itens = [
        {
            "task_id": int(d["task_id"]),
            "to_id": d.get("to_id"),
            "to_nome": d.get("to_nome"),
            "origem": d.get("origem") or "tarefa",
        }
        for d in (j.detalhe or [])
        if d.get("task_id") and d.get("reason") not in _REASON_OK
    ]
    if not itens:
        raise HTTPException(status_code=400, detail="Não há tarefas pendentes pra refazer nessa execução")

    novo = reatribuir_job.iniciar(team, itens, [], False, current_user)
    return {"job_id": novo, "total": len(itens), "status": reatribuir_job.status(novo)}


@router.get(
    "/reatribuir/jobs/{job_id}/excel",
    summary="Excel com o resultado de uma execução de reatribuição",
    dependencies=[_team],
)
def reatribuir_job_excel(job_id: str, team: str = Query(...), db: Session = Depends(get_db)):
    import io

    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    from app.models.performance import BalanceadorReatribuirJob

    j = db.get(BalanceadorReatribuirJob, job_id)
    if not j or j.team != team:
        raise HTTPException(status_code=404, detail="Execução não encontrada.")

    detalhe = reatribuir_job_detalhe(job_id, team=team, db=db)["tarefas"]

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")

    def _write_header(ws, cols):
        for idx, title in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=idx, value=title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"

    # ── Aba 1: Resumo ──
    ws = wb.active
    ws.title = "Resumo"
    resumo_cols = [
        "Execução", "Status", "Simulação", "Disparado por", "Iniciado em", "Terminado em",
        "Total", "Reatribuídas", "P/ manual", "Falhas",
    ]
    _write_header(ws, resumo_cols)
    ws.cell(row=2, column=1, value=j.id)
    ws.cell(row=2, column=2, value="Concluída" if j.status == "done" else j.status)
    ws.cell(row=2, column=3, value="Sim" if j.dry_run else "Não")
    ws.cell(row=2, column=4, value=j.criado_por_nome)
    ws.cell(row=2, column=5, value=j.iniciado_em.isoformat() if j.iniciado_em else None)
    ws.cell(row=2, column=6, value=j.terminado_em.isoformat() if j.terminado_em else None)
    ws.cell(row=2, column=7, value=j.total or 0)
    ws.cell(row=2, column=8, value=j.reatribuidas or 0)
    ws.cell(row=2, column=9, value=j.workflow_bloqueadas or 0)
    ws.cell(row=2, column=10, value=j.falhas or 0)
    for idx in range(1, len(resumo_cols) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 18

    # ── Aba 2: Tarefas (uma linha por tarefa, com motivo legível) ──
    ws2 = wb.create_sheet("Tarefas")
    tar_cols = ["Task ID", "Subtipo", "Pasta", "CNJ", "Destino", "Resultado", "HTTP"]
    _write_header(ws2, tar_cols)
    for i, d in enumerate(detalhe, start=2):
        ws2.cell(row=i, column=1, value=d.get("task_id"))
        ws2.cell(row=i, column=2, value=d.get("subtipo"))
        ws2.cell(row=i, column=3, value=d.get("pasta"))
        ws2.cell(row=i, column=4, value=d.get("cnj"))
        ws2.cell(row=i, column=5, value=d.get("to_nome"))
        ws2.cell(row=i, column=6, value=d.get("resultado"))
        ws2.cell(row=i, column=7, value=d.get("http"))
    for idx in range(1, len(tar_cols) + 1):
        ws2.column_dimensions[get_column_letter(idx)].width = 16
    ws2.column_dimensions[get_column_letter(2)].width = 44
    ws2.column_dimensions[get_column_letter(5)].width = 30
    ws2.column_dimensions[get_column_letter(6)].width = 32

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    stamp = (j.iniciado_em.strftime("%Y%m%d_%H%M") if j.iniciado_em else "exec")
    filename = f"redistribuicao_{team}_{stamp}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
